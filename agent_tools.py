import os
import re
import subprocess

import litellm
import json
import yaml
import openpyxl
import tempfile
import fnmatch
import logging
import textwrap
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Tuple, Callable, Optional, Set, Any
from google.adk.tools.tool_context import ToolContext
from utils.path_utils import normalize_patch_path, validate_patch_path
from utils.error_handler import format_path_error

logger = logging.getLogger(__name__)

ENABLE_REFLECTION = True
ENABLE_ROLLBACK = True
ENABLE_HISTORY_ENHANCEMENT = True
ENABLE_EXPERT_KNOWLEDGE = True

# Mechanism ablation mapping:
# - ENABLE_REFLECTION -> RSMC
# - ENABLE_ROLLBACK -> HSR / rollback
# - ENABLE_HISTORY_ENHANCEMENT -> ECRCL / history-constrained localization
# - ENABLE_EXPERT_KNOWLEDGE -> Few-shot RAG / expert knowledge injection

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

PROCESSED_PROJECTS_DIR = os.path.join(CURRENT_DIR, "process")
PROCESSED_PROJECTS_FILE = os.path.join(PROCESSED_PROJECTS_DIR, "project_processed.txt")
GLOBAL_CHAR_BUDGET = 280000  # 硬编码
max_lines = 2500  # 硬编码
_LATEST_BASIC_INFORMATION: Dict[str, Any] = {}


def extract_basic_information(raw_basic_information: Any) -> Dict[str, Any]:
    """
    Normalize `basic_information` into a structured dictionary.
    Accepts dict payloads or LLM-formatted strings that wrap a JSON object.
    """
    if isinstance(raw_basic_information, dict):
        data = dict(raw_basic_information)
    elif isinstance(raw_basic_information, str):
        data = {}
        json_match = re.search(r'(\{[\s\S]*\})', raw_basic_information)
        if json_match:
            try:
                data = json.loads(json_match.group(1))
            except Exception:
                data = {}
    else:
        data = {}

    if not isinstance(data, dict):
        data = {}

    if "project_name" not in data and data.get("project"):
        data["project_name"] = data["project"]

    for path_key in ["project_source_path", "project_config_path", "project_config_repo_path"]:
        if data.get(path_key):
            data[path_key] = os.path.abspath(data[path_key])

    return data


def _coerce_project_paths_with_basic_information(
        project_source_path: Optional[str],
        project_config_path: Optional[str],
        basic_info: Dict[str, Any]
) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Normalize upstream/downstream repository paths against structured basic_information.
    Prevents project root paths from degrading to parent directories like `process/project`.
    """
    expected_source_path = basic_info.get("project_source_path")
    expected_config_path = basic_info.get("project_config_path")
    expected_config_repo_path = basic_info.get("project_config_repo_path")

    if expected_source_path:
        normalized_source = os.path.abspath(project_source_path) if project_source_path else None
        expected_source_abs = os.path.abspath(expected_source_path)
        if normalized_source != expected_source_abs:
            project_source_path = expected_source_abs

    if expected_config_path:
        normalized_config = os.path.abspath(project_config_path) if project_config_path else None
        expected_config_abs = os.path.abspath(expected_config_path)
        if normalized_config != expected_config_abs:
            project_config_path = expected_config_abs

    project_config_repo_path = expected_config_repo_path
    if not project_config_repo_path and project_config_path:
        candidate_root = os.path.abspath(os.path.join(project_config_path, "..", ".."))
        if os.path.exists(os.path.join(candidate_root, ".git")):
            project_config_repo_path = candidate_root

    return project_source_path, project_config_path, project_config_repo_path


# =====================================================================
# 定位位置：fix_build_agent/agent_tools.py 中的 TraceLedgerManager 类
# 替换内容：引入 FILE_NAME 属性与自愈式活跃项目名追踪方法
# =====================================================================

class TraceLedgerManager:
    """
    Thread-safe Ledger State Tree persistence manager for project_repair_trace.json.
    Mediates all read/write interactions to keep the LLM decoupled from direct file system edits.
    """
    _active_project = "UNKNOWN"

    @classmethod
    def set_active_project(cls, project_name: str):
        cls._active_project = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()

    @classmethod
    def get_ledger_path(cls) -> str:
        return os.path.abspath(os.path.join(os.getcwd(), "project_repair_trace.json"))

    @classmethod
    def load_ledger(cls) -> dict:
        path = cls.get_ledger_path()
        if not os.path.exists(path):
            return {
                "project_name": cls._active_project,
                "archive_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "next_node_id": 0,
                "nodes": []
            }
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # 强制同步当前激活项目名
                data["project_name"] = cls._active_project
                if "next_node_id" not in data:
                    existing_ids = [n.get("node_id", -1) for n in data.get("nodes", []) if isinstance(n, dict)]
                    data["next_node_id"] = (max(existing_ids) + 1) if existing_ids else 0
                return data
        except Exception as e:
            return {"project_name": cls._active_project, "archive_date": "", "next_node_id": 0, "nodes": []}

    @classmethod
    def save_ledger(cls, data: dict) -> bool:
        path = cls.get_ledger_path()
        try:
            if "next_node_id" not in data:
                existing_ids = [n.get("node_id", -1) for n in data.get("nodes", []) if isinstance(n, dict)]
                data["next_node_id"] = (max(existing_ids) + 1) if existing_ids else 0
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            logger.error(f"Failed to save ledger: {e}")
            return False

    @classmethod
    def update_node_fields(cls, node_id: int, fields_dict: dict) -> bool:
        """
        Safely update arbitrary deep paths (using dot notation) of a specific node_id.
        """
        print(f"[DBG] update_node_fields node_id={node_id}, keys={list(fields_dict.keys())}")
        ledger = cls.load_ledger()
        target_node = None
        for node in ledger.get("nodes", []):
            if node.get("node_id") == node_id:
                target_node = node
                break

        if target_node is None:
            return False

        for key, value in fields_dict.items():
            parts = key.split('.')
            curr = target_node
            for part in parts[:-1]:
                if part not in curr:
                    curr[part] = {}
                curr = curr[part]
            curr[parts[-1]] = value

        print(f"[DBG] update_node_fields target_node_after={json.dumps(target_node, ensure_ascii=False)[:1200]}")
        return cls.save_ledger(ledger)

    @classmethod
    def get_node_by_id(cls, ledger: dict, node_id: int) -> Optional[dict]:
        for node in ledger.get("nodes", []):
            if node.get("node_id") == node_id:
                return node
        return None

    @classmethod
    def allocate_next_node_id(cls, ledger: dict) -> Tuple[int, dict]:
        if "next_node_id" not in ledger:
            existing_ids = [n.get("node_id", -1) for n in ledger.get("nodes", []) if isinstance(n, dict)]
            ledger["next_node_id"] = (max(existing_ids) + 1) if existing_ids else 0
        node_id = ledger["next_node_id"]
        ledger["next_node_id"] += 1
        return node_id, ledger

    @classmethod
    def get_git_head_sha(cls, repo_path: str) -> str:
        if not repo_path or not os.path.exists(repo_path):
            return "N/A"
        try:
            res = subprocess.run(["git", "-C", repo_path, "rev-parse", "HEAD"], capture_output=True, text=True,
                                 check=True)
            return res.stdout.strip()
        except Exception:
            return "N/A"

    @classmethod
    def get_patch_metrics(cls, repo_path: str) -> dict:
        """
        利用 git diff --numstat 获取最近一次提交的精确改动行数（HEAD~1 与 HEAD 之间）。
        """
        abs_path = os.path.abspath(repo_path)
        try:
            res = subprocess.run(
                ["git", "-C", abs_path, "diff", "--numstat", "HEAD~1", "HEAD"],
                capture_output=True, text=True, check=True
            )
            total_add = 0
            total_del = 0
            for line in res.stdout.splitlines():
                parts = line.split()
                if len(parts) >= 2:
                    total_add += int(parts[0])
                    total_del += int(parts[1])
            return {"Ladd": total_add, "Ldel": total_del}
        except Exception:
            # 兼容无 HEAD~1 的首个节点或异常情况
            return {"Ladd": 0, "Ldel": 0}


def _collect_parent_chain_node_ids(ledger: dict, start_node_id: int, limit: Optional[int] = None,
                                   include_self: bool = False) -> List[int]:
    chain_ids = []
    current_id = start_node_id
    if not include_self:
        start_node = TraceLedgerManager.get_node_by_id(ledger, start_node_id)
        if not start_node:
            return []
        current_id = start_node.get("parent_id", -1)

    visited = set()
    while current_id is not None and current_id != -1:
        if current_id in visited:
            break
        visited.add(current_id)
        node = TraceLedgerManager.get_node_by_id(ledger, current_id)
        if not node:
            break
        chain_ids.append(current_id)
        if limit is not None and len(chain_ids) >= limit:
            break
        current_id = node.get("parent_id", -1)

    return chain_ids


def _is_initial_round() -> bool:
    """
    辅助检测器：读取统一账本判断当前是否处于首轮（Node 0/Node 1 baseline 阶段）
    """
    ledger_path = "project_repair_trace.json"
    if os.path.exists(ledger_path):
        try:
            with open(ledger_path, 'r', encoding='utf-8') as lf:
                trace_data = json.load(lf)
                # 若账本中节点数 > 1（已建立 Node 1 且后续被 backfill 扩展），则说明进入了非初始轮的迭代
                if trace_data.get("nodes") and len(trace_data["nodes"]) > 1:
                    return False
        except Exception:
            pass
    return True


def update_trace_ledger(node_id: int, fields_dict: dict, repo_path: str = None,
                        tool_context: ToolContext = None) -> dict:
    """
    Secure backfilling tool for the Solution Applier Agent.
    Safely writes file diff metrics, active workspace, and Git SHAs into project_repair_trace.json.
    """
    try:
        basic_info = extract_basic_information(
            tool_context.session.state.get("basic_information")
            if tool_context and getattr(tool_context, "session", None)
            else _LATEST_BASIC_INFORMATION
        )
        target_file = fields_dict.get("action_and_intent.target_file", "")
        if target_file.startswith("process/project/"):
            repo_path = basic_info.get("project_source_path") or repo_path
        elif target_file.startswith("oss-fuzz/projects/"):
            repo_path = basic_info.get("project_config_repo_path") or repo_path

        print(
            "[DEBUG update_trace_ledger args] "
            f"node_id={node_id} | repo_path={repo_path} | "
            f"fields={json.dumps(fields_dict, ensure_ascii=False)[:1500]}"
        )
        ledger = TraceLedgerManager.load_ledger()
        existing_node = next((n for n in ledger.get("nodes", []) if n.get("node_id") == node_id), None)

        # Do not let placeholder SHAs overwrite real anchors already present in the ledger.
        if existing_node:
            for sha_key in ["git_sha_state.oss-fuzz_sha", "git_sha_state.project_sha"]:
                if sha_key in fields_dict:
                    incoming_val = fields_dict.get(sha_key)
                    existing_field = sha_key.split(".")[-1]
                    existing_val = existing_node.get("git_sha_state", {}).get(existing_field)
                    if incoming_val in [None, "", "N/A"] and existing_val not in [None, "", "N/A"]:
                        fields_dict.pop(sha_key, None)

        # 如果传入了 repo_path，在写入前自动提取物理 Git 指标进行覆盖，防御模型幻觉
        if repo_path and os.path.exists(repo_path):
            metrics = TraceLedgerManager.get_patch_metrics(repo_path)
            fields_dict["metrics.Ladd"] = metrics["Ladd"]
            fields_dict["metrics.Ldel"] = metrics["Ldel"]

        success = TraceLedgerManager.update_node_fields(node_id, fields_dict)
        if success:
            return {"status": "success", "message": f"Node {node_id} fields successfully backfilled."}
        else:
            return {"status": "error", "message": f"Failed to backfill Node {node_id} fields."}
    except Exception as e:
        return {"status": "error", "message": f"Exception occurred during backfilling: {str(e)}"}


def reclaim_path_permissions(path: str) -> bool:
    """
    【原子工具 1】高兼容性权限回收器。
    将指定路径的文件所有权安全夺回到宿主机当前用户，支持 Docker 强夺与宿主机 chmod 本地自愈双轨道。
    """
    import os
    import subprocess

    if not path or not os.path.exists(path):
        return True

    abs_path = os.path.abspath(path)
    uid, gid = os.getuid(), os.getgid()
    docker_ok = False

    # 1. 尝试使用 Docker alpine 容器强制回收权限（应对容器内 root 生成的顽固文件）
    try:
        result = subprocess.run([
            "docker", "run", "--rm", "-v", f"{abs_path}:/src",
            "alpine", "chown", "-R", f"{uid}:{gid}", "/src"
        ], capture_output=True, text=True, timeout=15, check=False)
        if result.returncode == 0:
            docker_ok = True
    except Exception:
        pass

    # 2. 如果 Docker 不可用/超时，自动降级至 Host 本地 native 赋权自愈
    if not docker_ok:
        try:
            # 强行赋予当前宿主机用户读、写、执行权限
            subprocess.run(["chmod", "-R", "u+rwX", abs_path], capture_output=True, check=False)
        except Exception:
            pass

    return True


def safe_delete_path(path: str) -> bool:
    """
    【原子工具 2】安全物理删除器。
    在删除任何文件或目录前，先自动夺回权限，防止 PermissionError 导致流程崩溃。
    """
    import os
    import shutil

    if not path or not os.path.exists(path):
        return True

    abs_path = os.path.abspath(path)

    # 1. 先安全夺回权限
    reclaim_path_permissions(abs_path)

    # 2. 物理彻底删除
    try:
        if os.path.isdir(abs_path):
            shutil.rmtree(abs_path, ignore_errors=True)
        else:
            os.remove(abs_path)
        return True
    except Exception as e:
        print(f"--- [Warning] Failed to physically remove {abs_path}: {e} ---")
        return False


def _safe_path_wrapper(*d_args, **d_kwargs):
    """
    完全自适应型安全路径拦截包装装饰器。
    无缝兼容以下四种主流 Python 装饰器声明语法，绝不产生编译期或运行期类型报错：
    1. 无参直接修饰：  @_safe_path_wrapper
    2. 无参括号修饰：  @_safe_path_wrapper()
    3. 有参位置修饰：  @_safe_path_wrapper("save_file_tree_shallow")
    4. 有参关键字修饰：@_safe_path_wrapper(operation_name="read_file_content")
    """
    import functools
    import os
    from typing import Callable
    from google.adk.tools.tool_context import ToolContext

    # 1. 编译期解析有参传入的操作名称（兼容位置/关键字参数）
    op_name = None
    if d_args and isinstance(d_args[0], str):
        op_name = d_args[0]
    elif "operation_name" in d_kwargs:
        op_name = d_kwargs["operation_name"]

    def decorator(func: Callable) -> Callable:
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            # 自动提取操作名（有参优先，无参则自动使用被修饰函数名）
            operation_name = op_name if op_name else func.__name__

            # 2. 安全提取 ToolContext（深度兼容位置参数与关键字参数）
            tool_context = None
            for arg in args:
                if isinstance(arg, ToolContext):
                    tool_context = arg
                    break
            if not tool_context:
                tool_context = kwargs.get("tool_context")

            # 3. HSR 物理与认知回退空跑保护机制 (Bypass Protection)
            if tool_context and (
                    tool_context.session.state.get("rollback_triggered") or
                    tool_context.session.state.get("should_rollback")
            ):
                print(f"--- ⚠️ [STATE BYPASS] {operation_name} bypassed due to active rollback state. ---")
                return {
                    "status": "skipped",
                    "message": f"Rollback triggered. Bypassing further {operation_name} edits in this round."
                }

            # 4. 强类型安全路径参数抓取 ( args[0] 必须是 str，防止把 args 里的 ToolContext 误当路径)
            path_arg = (
                    kwargs.get('file_path') or
                    kwargs.get('directory_path') or
                    kwargs.get('source_path') or
                    kwargs.get('destination_path') or
                    kwargs.get('dir_path') or
                    kwargs.get('solution_file_path') or
                    (args[0] if args and isinstance(args[0], str) else None)
            )

            if not path_arg:
                # 辅助性零路径探测：若函数本身确实无路径参数传入，则直接执行并解脱校验
                return func(*args, **kwargs)

            # 获取项目基准路径
            base_dir = kwargs.get('base_dir', os.environ.get('PROJECT_ROOT', os.getcwd()))
            strict_mode = kwargs.get('strict_mode', True)

            # 导入白名单验证逻辑与标准路径错误引导
            from utils.path_utils import normalize_patch_path, validate_patch_path
            from utils.error_handler import format_path_error

            normalized = normalize_patch_path(path_arg, base_dir)
            if strict_mode and not validate_patch_path(normalized, strict=True):
                return {
                    "status": "error",
                    "message": format_path_error(
                        original_path=path_arg,
                        normalized_path=normalized,
                        base_dir=base_dir,
                        validation_passed=False,
                        extra_info={'operation': operation_name}
                    )
                }

            # 规范化 kwargs 里的路径参数
            path_keys = ['file_path', 'directory_path', 'source_path', 'destination_path', 'dir_path',
                         'solution_file_path']
            for key in path_keys:
                if key in kwargs:
                    kwargs[key] = normalized

            # 规范化 args 里的位置路径参数（只有在确定 args[0] 物理匹配时才重写）
            new_args = list(args)
            if args and isinstance(args[0], str) and args[0] == path_arg:
                new_args[0] = normalized

            return func(*new_args, **kwargs)

        return wrapper

    # 核心自适应分流判定
    if d_args and callable(d_args[0]):
        func = d_args[0]
        return decorator(func)
    else:
        return decorator


def get_verified_git_sha(repo_path: str, retries: int = 3) -> str:
    """带审计重试机制的 SHA 获取，确保 Git 节点确实可用"""
    import subprocess
    import time
    for i in range(retries):
        # 1. 检查目录是否存在
        if not os.path.exists(repo_path):
            return "N/A"
        # 2. 尝试获取 HEAD SHA
        res = subprocess.run(["git", "-C", repo_path, "rev-parse", "HEAD"],
                             capture_output=True, text=True)
        if res.returncode == 0:
            sha = res.stdout.strip()
            print(f"--- [Audit] SHA valid: {sha} (Attempt {i + 1}) ---")
            return sha

        # 3. 如果失败且确认未初始化，尝试进行补救性 init
        if not os.path.exists(os.path.join(repo_path, ".git")):
            print(f"--- [Audit] Repo not initialized at {repo_path}, triggering init ---")
            manage_git_state(repo_path, "init")

        print(f"--- [Audit] SHA invalid at {repo_path}, retrying... ({i + 1}/{retries}) ---")
        time.sleep(1)

    print(f"--- [Audit] CRITICAL: SHA invalid after {retries} retries in {repo_path} ---")
    return "N/A"



def get_absolute_host_path(relative_path: str) -> str:
    """
    Host-Path Normalization Engine:
    Translates Agent-provided paths into host-machine absolute paths.

    Logic:
    1. Strips container-style prefixes like /workspace/, /src/, /work/.
    2. Maps paths to either 'process/project/' or 'oss-fuzz/'.
    3. Guarantees return of a valid absolute host path.
    """
    import os

    workspace_root = os.getcwd()
    if not relative_path:
        return workspace_root

    normalized_input = os.path.normpath(str(relative_path))
    if os.path.isabs(normalized_input) and os.path.exists(normalized_input):
        return normalized_input

    path = normalized_input.lstrip('/')

    # 1. Path Stripping (Removing hallucinatory container prefixes)
    for prefix in ['workspace/', 'src/', 'work/', 'fix_build_agent/']:
        if path.startswith(prefix):
            path = path[len(prefix):]
            break

    path = os.path.normpath(path)

    # 2. Structured Path Remapping
    if path.startswith('oss-fuzz/') or path.startswith('process/project/'):
        return os.path.abspath(os.path.join(workspace_root, path))

    if path.startswith('projects/'):
        return os.path.abspath(os.path.join(workspace_root, 'oss-fuzz', path))

    path_parts = path.split(os.sep)
    first_part = path_parts[0] if path_parts else ''

    candidate_paths = []
    if first_part:
        candidate_paths.extend([
            os.path.join(workspace_root, 'process', 'project', path),
            os.path.join(workspace_root, 'oss-fuzz', 'projects', path),
            os.path.join(workspace_root, 'oss-fuzz', path),
        ])

    for candidate in candidate_paths:
        if os.path.exists(candidate):
            return os.path.abspath(candidate)

    # 3. Default to workspace root resolution
    return os.path.abspath(os.path.join(workspace_root, path))


@_safe_path_wrapper
def apply_patch(solution_file_path: str, **kwargs) -> dict:
    import difflib
    base_dir = kwargs.get('base_dir', os.getcwd())
    if not os.path.exists(solution_file_path):
        return {"status": "error", "message": "Solution file not found."}
    try:
        with open(solution_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        patch_blocks = content.split('---=== FILE ===---')[1:]
        applied_count = 0
        errors = []

        for block in patch_blocks:
            parts = block.split('---=== ORIGINAL ===---')
            original_target = parts[0].strip()
            content_parts = parts[1].split('---=== REPLACEMENT ===---')
            original_block = content_parts[0].strip("\n\r")
            replacement_block = content_parts[1].strip("\n\r")

            file_path = os.path.normpath(os.path.join(base_dir, original_target)) if not os.path.isabs(
                original_target) else original_target
            if not os.path.exists(file_path):
                errors.append(f"File not found: {original_target}")
                continue

            with open(file_path, 'r', encoding='utf-8') as f:
                file_content = f.read()

            if replacement_block in file_content:
                applied_count += 1
                continue

            norm_repl = re.sub(r'\s+', ' ', replacement_block).strip()
            norm_file = re.sub(r'\s+', ' ', file_content).strip()
            if norm_repl in norm_file:
                applied_count += 1
                continue

            if original_block in file_content:
                new_content = file_content.replace(original_block, replacement_block, 1)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)
                applied_count += 1
                continue

            norm_orig = re.sub(r'\s+', ' ', original_block).strip()
            norm_file = re.sub(r'\s+', ' ', file_content).strip()
            if norm_orig in norm_file:
                file_lines = file_content.splitlines()
                orig_lines = original_block.splitlines()
                matched_idx = -1
                for i in range(len(file_lines) - len(orig_lines) + 1):
                    window = "\n".join(file_lines[i:i + len(orig_lines)])
                    if re.sub(r'\s+', ' ', window).strip() == norm_orig:
                        matched_idx = i
                        break
                if matched_idx != -1:
                    new_lines = file_lines[:matched_idx] + replacement_block.splitlines() + file_lines[
                        matched_idx + len(
                            orig_lines):]
                    with open(file_path, 'w', encoding='utf-8') as f:
                        f.write("\n".join(new_lines))
                    applied_count += 1
                    continue

            lines = file_content.splitlines()
            search_anchor = original_block.splitlines()[0].strip()
            matches = difflib.get_close_matches(search_anchor, lines, n=1, cutoff=0.3)
            ctx = "Unknown context"
            if matches:
                idx = lines.index(matches[0])
                ctx = "\n".join(lines[max(0, idx - 5):min(len(lines), idx + 10)])
            errors.append(f"MATCH FAILED for {original_target}.\n### ACTUAL CONTENT AROUND TARGET AREA ###\n{ctx}")

            # 统计所有块的增删行数（在 for block 循环结束后执行）
        total_lines_changed = 0
        for stat_block in patch_blocks:
            stat_parts = stat_block.split('---=== ORIGINAL ===---')
            if len(stat_parts) < 2:
                continue
            stat_content_parts = stat_parts[1].split('---=== REPLACEMENT ===---')
            if len(stat_content_parts) < 2:
                continue
            orig_line_count = len(stat_content_parts[0].strip("\n\r").splitlines())
            repl_line_count = len(stat_content_parts[1].strip("\n\r").splitlines())
            total_lines_changed += max(orig_line_count, repl_line_count)

        return {
            "status": "success" if not errors else "error",
            "modified_files_count": applied_count,
            "modified_lines_count": total_lines_changed,
            "errors": errors
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def manage_git_state(path: str, action: str, message: str = "", commit_sha: str = "") -> Dict:
    """
    Manages the Git state tree with logical fencing and physical auditing.
    Supports: init, commit, rollback, status, log, fetch.
    """
    physical_path = get_absolute_host_path(path)

    if not os.path.exists(physical_path):
        return {'status': 'error', 'message': f'Resolved path {physical_path} does not exist.'}
    print(f"--- Tool: manage_git_state | Action: {action} | Path: {path} ---")

    if not os.path.exists(path):
        return {"status": "error", "message": f"Path {path} does not exist."}

    abs_path = os.path.abspath(path)
    framework_root = os.path.dirname(os.path.abspath(__file__))
    if abs_path == framework_root:
        return {"status": "error",
                "message": "CRITICAL: Security Violation. Operations on Agent Framework root are blocked."}

    original_cwd = os.getcwd()
    try:
        # 1. 🔑 物理环境权限自愈（单行调用原子工具，替换原有繁琐的 Docker 代码块）
        if action in ["init", "commit", "rollback"]:
            reclaim_path_permissions(abs_path)

        os.chdir(abs_path)

        # 2. 基础配置初始化
        if action in ["init", "commit"]:
            if not os.path.exists(".git"):
                subprocess.run(["git", "init"], check=True, capture_output=True)
            subprocess.run(["git", "config", "user.email", "agent@oss-fuzz-repair.com"], check=True)
            subprocess.run(["git", "config", "user.name", "Repair Agent"], check=True)

        # 3. 分支逻辑处理
        if action == "init":
            subprocess.run(["git", "commit", "--allow-empty", "-m", "[BASELINE] Initial state"], check=True)
            subprocess.run(["git", "add", "."], check=True)
            has_commit = subprocess.run(["git", "rev-parse", "HEAD"], capture_output=True).returncode == 0
            if not has_commit:
                subprocess.run(["git", "commit", "-m", "[BASELINE] Initial experiment state"], check=True,
                               capture_output=True)
            return {"status": "success", "message": f"Git initialized at Baseline in {path}"}


        elif action == "commit":
            subprocess.run(["git", "add", "."], check=True)
            full_message = f"[AGENT_FIX] {message}"
            subprocess.run(["git", "commit", "--allow-empty", "-m", full_message], capture_output=True, text=True, check=True)
            sha = subprocess.run(["git", "rev-parse", "HEAD"], capture_output=True, text=True).stdout.strip()
            return {"status": "success", "sha": sha, "message": f"State saved: {full_message}"}

        elif action == "rollback":
            # 统计带有 [AGENT_FIX] 标记的提交数量作为配额
            res = subprocess.run(["git", "log", "--grep=\\[AGENT_FIX\\]", "--oneline"], capture_output=True, text=True)
            quota = len([l for l in res.stdout.split('\n') if l.strip()])

            if quota <= 0:
                return {
                    "status": "error",
                    "message": "Already at the Initial Baseline. No further rollback possible."
                }

            target = commit_sha if commit_sha else "HEAD~1"
            subprocess.run(["git", "reset", "--hard", target], check=True, capture_output=True)
            subprocess.run(["git", "clean", "-fxd"], check=True, capture_output=True)
            return {"status": "success", "message": f"Rolled back to {target}. Remaining Fixes: {quota - 1}"}

        elif action == "status":
            res = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True)
            return {"status": "success", "stdout": res.stdout, "message": "Retrieved git status."}

        elif action == "log":
            # 支持将 message 解析为附加参数，例如 "-n 5"
            log_args = message.split() if message else ["-n", "5", "--oneline"]
            cmd = ["git", "log"] + log_args
            res = subprocess.run(cmd, capture_output=True, text=True)
            return {"status": "success", "stdout": res.stdout, "message": f"Executed: {' '.join(cmd)}"}

        elif action == "fetch":
            # 处理远程拉取请求
            fetch_args = message.split() if message else ["origin"]
            cmd = ["git", "fetch"] + fetch_args
            res = subprocess.run(cmd, capture_output=True, text=True)
            return {"status": "success", "stdout": res.stdout, "message": "Fetch completed."}

        else:
            return {"status": "error", "message": f"Action '{action}' is not implemented."}

    except Exception as e:
        return {"status": "error", "message": f"Git Intervention Failed: {str(e)}"}
    finally:
        os.chdir(original_cwd)


def commit_workspace_snapshots(project_source_path: str, project_config_path: str, attempt_id: int) -> Dict[str, str]:
    """
    Create synchronized Git snapshot commits for both upstream and downstream repositories.
    Each workspace always receives a new commit SHA via `git add . && git commit --allow-empty -m ...`.
    """
    basic_info = extract_basic_information(_LATEST_BASIC_INFORMATION)
    project_source_path, project_config_path, downstream_repo_path = _coerce_project_paths_with_basic_information(
        project_source_path,
        project_config_path,
        basic_info
    )

    downstream_repo_path = downstream_repo_path or project_config_path
    if downstream_repo_path and not os.path.exists(os.path.join(downstream_repo_path, ".git")):
        candidate_root = os.path.abspath(os.path.join(downstream_repo_path, "..", ".."))
        if os.path.exists(os.path.join(candidate_root, ".git")):
            downstream_repo_path = candidate_root
    print(
        "[DEBUG commit_workspace_snapshots args] "
        f"project_source_path={project_source_path} | "
        f"project_config_path={project_config_path} | "
        f"downstream_repo_path={downstream_repo_path} | "
        f"attempt_id={attempt_id}"
    )

    upstream_res = manage_git_state(
        path=project_source_path,
        action="commit",
        message=f"Applied fix for attempt {attempt_id} (UPSTREAM SNAPSHOT)"
    )
    if upstream_res.get("status") != "success" or not upstream_res.get("sha"):
        return {
            "status": "error",
            "message": f"Failed to create upstream snapshot: {upstream_res.get('message', 'Unknown error')}"
        }

    downstream_res = manage_git_state(
        path=downstream_repo_path,
        action="commit",
        message=f"Applied fix for attempt {attempt_id} (DOWNSTREAM SNAPSHOT)"
    )
    if downstream_res.get("status") != "success" or not downstream_res.get("sha"):
        return {
            "status": "error",
            "message": f"Failed to create downstream snapshot: {downstream_res.get('message', 'Unknown error')}",
            "project_sha": upstream_res.get("sha")
        }

    return {
        "status": "success",
        "project_sha": upstream_res.get("sha"),
        "oss_fuzz_sha": downstream_res.get("sha"),
        "message": "Synchronized snapshots committed for both workspaces."
    }


# run_fuzz_and_collect_log_agent tools

def run_fuzz_build_and_validate(
        project_name: str,
        oss_fuzz_path: str,
        sanitizer: str,
        engine: str,
        architecture: str,
        mount_path: Optional[str] = None,
        verbose_step6: bool = False,
        verbose_build: bool = True,
        tool_context: ToolContext = None
) -> dict:
    """
    Build and validate fuzzers using official OSS-Fuzz infrastructure.
    Success Criteria: Step 2 (check_build) must PASS. All other steps are reference items.
    """

    import stat
    import subprocess
    import time
    import os
    import sys
    import signal
    import select  # 用于非阻塞读取
    import re  # 用于进度正则匹配

    raw_basic_information = None
    if tool_context and getattr(tool_context, "session", None):
        raw_basic_information = tool_context.session.state.get("basic_information")
    basic_info = extract_basic_information(raw_basic_information or _LATEST_BASIC_INFORMATION)

    project_name = basic_info.get("project_name") or project_name
    sanitizer = basic_info.get("sanitizer") or sanitizer
    engine = basic_info.get("engine") or engine
    architecture = basic_info.get("architecture") or architecture
    if basic_info.get("project_source_path"):
        mount_path = basic_info.get("project_source_path")
    if basic_info.get("project_config_repo_path"):
        oss_fuzz_path = basic_info.get("project_config_repo_path")
    elif basic_info.get("project_config_path"):
        oss_fuzz_path = os.path.abspath(os.path.join(basic_info.get("project_config_path"), "..", ".."))

    print(
        "[DEBUG run_fuzz_build_and_validate args] "
        f"project_name={project_name} | "
        f"oss_fuzz_path={oss_fuzz_path} | "
        f"sanitizer={sanitizer} | "
        f"engine={engine} | "
        f"architecture={architecture} | "
        f"mount_path={mount_path}"
    )

    _cleanup_environment(oss_fuzz_path, project_name)
    print(f"[*] Comprehensive Pre-build cleanup completed.")

    LOG_DIR = "fuzz_build_log_file"
    LOG_FILE_PATH = os.path.join(LOG_DIR, "fuzz_build_log.txt")
    os.makedirs(LOG_DIR, exist_ok=True)

    report = {
        "step_1_official_list": "pending",
        "step_2_infra_compliance": "pending",
        "step_3_sanitizer_injected": "pending",
        "step_4_engine_control": "pending",
        "step_5_logic_linkage": "pending",
        "step_6_runtime_stability": "pending"
    }

    def build_summary_table() -> str:
        summary = "\n" + "=" * 50 + "\n--- VALIDATION SUMMARY\n" + "-" * 50 + "\n"
        for i, (k, v) in enumerate(report.items(), 1):
            marker = "[MANDATORY]" if i == 2 else "[REFERENCE]"
            summary += f"Step {i:<4} {marker:<12} | {v}\n"
        summary += "=" * 50 + "\n"
        return summary

    def write_log_artifact(base_log: str, result_line: str) -> None:
        with open(LOG_FILE_PATH, "w", encoding="utf-8") as f:
            f.write(base_log)
            f.write(build_summary_table())
            f.write(f"\n{result_line}")

    # =========================================================================
    # 内部辅助过滤函数（对应 test_all.py 中的合法 Fuzzer 识别逻辑）
    # =========================================================================
    def is_elf(filepath: str) -> bool:
        """判断是否为 ELF 格式二进制文件"""
        try:
            result = subprocess.run(
                ['file', filepath],
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                check=False
            )
            if b'ELF' in result.stdout:
                return True
        except Exception:
            pass
        # 兜底：如果系统没有安装 file 命令，直接读取文件头部魔数进行基础判断
        try:
            with open(filepath, 'rb') as f:
                return f.read(4) == b'\x7fELF'
        except Exception:
            return False

    def is_shell_script(filepath: str) -> bool:
        """判断是否为 shell 脚本"""
        try:
            result = subprocess.run(
                ['file', filepath],
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
                check=False
            )
            return b'shell script' in result.stdout
        except Exception:
            return False

    def find_local_fuzz_targets(directory: str, target_engine: str) -> list:
        """基于 test_all.py 标准过滤机制定位合法构建产物"""
        fuzz_targets = []
        if not os.path.exists(directory):
            return fuzz_targets

        executable_mask = stat.S_IEXEC | stat.S_IXGRP | stat.S_IXOTH

        for filename in os.listdir(directory):
            path = os.path.join(directory, filename)

            # ---- 第一层过滤：物理属性过滤 (Structural Filter) ----
            # 1. 排除特定辅助工具与非 Fuzzer 产物
            if filename == 'llvm-symbolizer':
                continue
            if filename.startswith('afl-'):
                continue
            if filename.startswith('jazzer_'):
                continue
            if filename == 'centipede':
                continue

            # 2. 必须是文件
            if not os.path.isfile(path):
                continue

            # 3. 必须具备可执行权限
            try:
                if not (os.stat(path).st_mode & executable_mask):
                    continue
            except Exception:
                continue

            # 4. 必须是 ELF 二进制或 Shell 脚本包装器
            if not is_elf(path) and not is_shell_script(path):
                continue

            # ---- 第二层过滤：符号合规性过滤 (Symbol Filter) ----
            if target_engine not in {'none', 'wycheproof'}:
                try:
                    with open(path, 'rb') as file_handle:
                        binary_contents = file_handle.read()
                        if b'LLVMFuzzerTestOneInput' not in binary_contents:
                            continue
                except Exception:
                    continue

            fuzz_targets.append(filename)
        return fuzz_targets

    # =========================================================================

    try:
        helper_path = os.path.join(oss_fuzz_path, "infra/helper.py")

        # --- Phase 1: Physical Build ---
        build_cmd = ["python3", helper_path, "build_fuzzers"]
        # 强制始终挂载 project_source_path
        build_cmd.extend([project_name, mount_path])
        build_cmd.extend(["--sanitizer", sanitizer, "--engine", engine, "--architecture", architecture])

        build_start = time.time()
        build_timeout = 5400  # 构建超时上限设定为 90 分钟（5400 秒），不影响正常构建结束

        process = subprocess.Popen(
            build_cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, cwd=oss_fuzz_path
        )
        full_log = []

        try:
            while True:
                if time.time() - build_start > build_timeout:
                    raise subprocess.TimeoutExpired(build_cmd, build_timeout)
                line = process.stdout.readline()
                if not line:
                    if process.poll() is not None:
                        break
                    time.sleep(0.05)
                    continue
                if verbose_build:
                    print(line, end='', flush=True)
                full_log.append(line)
            process.wait(timeout=15)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait()
            final_log = "".join(full_log) + f"\n\nRESULT: failed (compilation timeout after {build_timeout}s)"
            write_log_artifact("".join(full_log), f"RESULT: failed (compilation timeout after {build_timeout}s)")
            return {"status": "error", "message": "Compilation timed out", "validation_report": report}

        final_log = "".join(full_log)

        # 编译失败检测：仅依据构建进程退出码判定，避免日志关键词误伤后续 Step 2 成功场景
        if process.returncode != 0:
            write_log_artifact(final_log, "RESULT: failed (compilation error)")
            return {"status": "error", "message": "Compilation failed", "validation_report": report}

        # --- Phase 2: Deep Validation ---
        print(f"\n--- [Phase 2] Deep Validation (Official Suite) ---")

        # 🔑 建立独立验证阶段计时锁，上限调整为 20 分钟 (1200.0 秒)，不影响正常验证结束
        validation_start_time = time.time()
        validation_timeout = 1200.0

        def check_validation_limit(cmd_info):
            elapsed = time.time() - validation_start_time
            if elapsed >= validation_timeout:
                raise subprocess.TimeoutExpired(cmd_info, validation_timeout)
            return validation_timeout - elapsed

        # =========================================================================
        # Step 1: 官方产物识别 (参考项)
        # =========================================================================
        _ = check_validation_limit("list_fuzzers")
        out_dir = os.path.join(oss_fuzz_path, "build", "out", project_name)

        # 使用高保真度本地过滤逻辑
        targets = find_local_fuzz_targets(out_dir, engine)

        # 在日志中全量输出检测到的所有合规的模糊构建产物
        print(
            f"[*] Detected {len(targets)} compliant fuzz target(s) in Step 1: {', '.join(targets) if targets else 'None'}")

        primary_target = None
        if targets:
            primary_target = targets[0]
            report["step_1_official_list"] = f"pass: {len(targets)} target(s) (primary: {primary_target})"
        else:
            report["step_1_official_list"] = "fail: no recognized fuzzers"

        # =========================================================================
        # Step 2: 基础设施合规性 (唯一强制通过项)
        # =========================================================================
        rem_t = check_validation_limit("check_build")
        check_cmd = [
            "python3", helper_path, "check_build", project_name,
            "--sanitizer", sanitizer,
            "--engine", engine,
            "--architecture", architecture
        ]
        try:
            check_res = subprocess.run(check_cmd, capture_output=True, text=True, timeout=min(300, rem_t),
                                       cwd=oss_fuzz_path)
            report[
                "step_2_infra_compliance"] = "pass" if check_res.returncode == 0 else f"fail: {check_res.stderr.strip()[:200]}"
        except subprocess.TimeoutExpired:
            report["step_2_infra_compliance"] = "fail: check_build timeout"
        except Exception as e:
            report["step_2_infra_compliance"] = f"fail: {str(e)}"

        # Step 3-5: 参考项审计 (nm 符号分析 - 参考项)
        if primary_target:
            target_path = os.path.join(oss_fuzz_path, "build", "out", project_name, primary_target)
            if os.path.exists(target_path):
                rem_t = check_validation_limit("nm_check")
                try:
                    nm_res = subprocess.run(['nm', target_path], capture_output=True, text=True,
                                            timeout=min(30, rem_t),
                                            errors='ignore')
                    nm_stdout = nm_res.stdout
                except Exception:
                    rem_t = check_validation_limit("nm_check_shell")
                    nm_res = subprocess.run(
                        ["python3", helper_path, "shell", project_name, "-c", f"nm /out/{primary_target}"],
                        capture_output=True, text=True, timeout=min(60, rem_t), errors='ignore'
                    )
                    nm_stdout = nm_res.stdout

                report["step_3_sanitizer_injected"] = "pass" if "__asan" in nm_stdout else "warning: missing asan"
                report["step_4_engine_control"] = "pass" if (
                        "LLVMFuzzerRunDriver" in nm_stdout or "__afl_" in nm_stdout) else "warning: engine symbols"
                report["step_5_logic_linkage"] = "pass" if _auto_discover_project_symbols_from_content(nm_stdout,
                                                                                                       project_name) else "warning: logic linkage"
            else:
                for s in ["step_3_sanitizer_injected", "step_4_engine_control", "step_5_logic_linkage"]:
                    report[s] = "skip: binary not accessible"
        else:
            for s in ["step_3_sanitizer_injected", "step_4_engine_control", "step_5_logic_linkage"]:
                report[s] = "skip: no primary target"

        # =========================================================================
        # Step 6: 压力测试稳定性 (参考项)
        # =========================================================================
        if primary_target and "pass" in str(report["step_2_infra_compliance"]).lower():
            if verbose_step6:  # ✅ 仅在开启 verbose 模式时输出 Header
                print(f"[*] Starting 35s stability test for: {primary_target}")
            run_cmd = [sys.executable, helper_path, "run_fuzzer", "--engine", engine, "--sanitizer", sanitizer,
                       project_name, primary_target]

            rem_t = check_validation_limit("run_fuzzer")

            # 开启新进程组
            stability_proc = subprocess.Popen(
                run_cmd, cwd=oss_fuzz_path, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, preexec_fn=os.setsid
            )

            start_time = time.time()
            log_lines = []
            timed_out = False

            try:
                while True:
                    check_validation_limit("run_fuzzer_runtime")

                    elapsed = time.time() - start_time
                    if elapsed >= 35.0:
                        timed_out = True
                        break

                    remaining_time = max(0.1, 35.0 - elapsed)
                    rlist, _, _ = select.select([stability_proc.stdout], [], [], min(remaining_time, 0.5))

                    if stability_proc.stdout in rlist:
                        line = stability_proc.stdout.readline()
                        if not line: break
                        if verbose_step6:  # ✅ 仅在开启时实时打印模糊测试进度
                            print(line, end='', flush=True)
                        log_lines.append(line)
                    else:
                        if stability_proc.poll() is not None:
                            break
            finally:
                # 🔑 2. 安全优雅释放逻辑
                pgid = None
                try:
                    pgid = os.getpgid(stability_proc.pid)
                except Exception:
                    pass

                if pgid:
                    try:
                        # 优先发送 SIGTERM，允许 helper.py 在退出前调用 docker 释放容器卷锁
                        os.killpg(pgid, signal.SIGTERM)
                    except Exception:
                        pass

                # 阻塞回收进程状态
                try:
                    stability_proc.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    if pgid:
                        try:
                            # 如果 5 秒后仍未释放，再使用 SIGKILL 强制物理杀死
                            os.killpg(pgid, signal.SIGKILL)
                        except Exception:
                            pass
                    stability_proc.wait()

                if verbose_step6:  # ✅ 仅在开启时打印结束信息
                    print("[*] Stability test exited. Triggering comprehensive post-validation cleanup...")
                _cleanup_environment(oss_fuzz_path, project_name)

            # 日志文本整合与退出码转换
            log_content = "".join(log_lines)
            exit_code = 124 if timed_out else stability_proc.returncode
            if exit_code is None:
                exit_code = 124

            # ---- 成功特征检测与失败规则匹配 ----
            # 🌟 修复关键点：拓宽正则表达式，包含 AFL++ 和其它引擎进度日志的标志词 (如 exec speed, corpus count, cycles done, etc.)
            progress_pattern = r'(exec/s:|cov:|corp:|exec speed|corpus count|cycles done|execs/sec|active execution rate)'
            has_progress = bool(re.search(progress_pattern, log_content, re.IGNORECASE))
            is_success_6 = False
            success_reason = ""

            # A. 优先执行显式成功逻辑判定 (Success Logic)
            # 成功情况 1：进程超时正常退出且日志中存在关键变异进度
            if exit_code == 124 and has_progress:
                is_success_6 = True
                success_reason = "pass: Time-limited run completed successfully."
            # 成功情况 2：引擎平稳退出，且日志显示完成
            elif exit_code == 0 and any(kw in log_content for kw in ["Done", "fuzzing finished"]):
                is_success_6 = True
                success_reason = "pass: Finished normally."

            # B. 若不满足显式成功，执行失败判定过滤；若非检测到的失败条件，则依然判定为成功
            if not is_success_6:
                is_failed_6 = False
                fail_reason = ""

                # 失败条件 B-1: 启动即崩溃/运行时 Crash (严重)
                if "SUMMARY:" in log_content or "AddressSanitizer" in log_content or "Segmentation fault" in log_content:
                    is_failed_6 = True
                    fail_reason = "fail: RUNTIME_CRASH"

                # 失败条件 B-2: 配置/路径/环境不匹配 (启动失败)
                elif exit_code in [1, 127] or any(k in log_content for k in
                                                  ["error while loading shared libraries", "undefined reference",
                                                   "Usage:"]):
                    is_failed_6 = True
                    fail_reason = "fail: CONFIG_ERROR"

                # 失败条件 B-3: 伪运行 (Dead/Frozen)
                elif exit_code == 124 and not has_progress:
                    is_failed_6 = True
                    fail_reason = "fail: DEAD_PROCESS"

                # 失败条件 B-4: 其它判定失败的非正常退出码（且排除 0 和 124）
                elif exit_code != 0 and exit_code != 124:
                    is_failed_6 = True
                    fail_reason = f"fail: Exit code {exit_code}"

                # 结论：如果不符合以上任何失败条件，依然判定为成功
                if not is_failed_6:
                    report["step_6_runtime_stability"] = "pass: Default success (No failure criteria matched)"
                else:
                    report["step_6_runtime_stability"] = fail_reason
            else:
                report["step_6_runtime_stability"] = success_reason
        else:
            report["step_6_runtime_stability"] = "fail: skipped"

        # --- 最终判定逻辑 (🌟 当前仅以 Step 2 check_build 作为唯一强约束通过项) ---
        is_success = "pass" in str(report["step_2_infra_compliance"]).lower()

        summary_table = build_summary_table()
        print(summary_table)

        # 写入物理日志
        write_log_artifact(final_log, f"RESULT: {'success' if is_success else 'failed'}")

        return {
            "status": "success" if is_success else "error",
            "message": f"Validation {'PASSED' if is_success else 'FAILED'}",
            "validation_report": report
        }

    except subprocess.TimeoutExpired as e:
        print(f"\n[⚠️ TIMEOUT] Validation phase exceeded limit. Aborting...")
        write_log_artifact("Validation phase timed out.\n", "RESULT: failed (compilation error)")
        return {"status": "error", "message": "Compilation failed", "validation_report": report}

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        write_log_artifact(f"Exception during validation:\n{str(e)}\n{tb}\n", "RESULT: failed (compilation error)")
        return {"status": "error", "message": str(e), "validation_report": report}


# rollback_agent tools

def cbsc_classify_log(log_path: str = None, dpseek_key: str = None) -> dict:
    """
    Stateless Cascade Build Stage Classifier (CBSC).
    Runs Phase 1 & 2 Reverse-order Cascade Search, and falls back to
    Phase 3 LLM Arbitration under ambiguous or silent exit scenarios.
    """
    # 日志路径优先级：实时构建日志 > 历史错误日志 > 默认实时日志
    real_time_log = "fuzz_build_log_file/fuzz_build_log.txt"
    if os.path.exists(real_time_log):
        log_path = real_time_log
    elif _is_initial_round():
        historical_dir = "build_error_log"
        found_historical = False
        if os.path.exists(historical_dir):
            for root, _, files in os.walk(historical_dir):
                for f in files:
                    if f.endswith(".txt") and "error" in f.lower():
                        log_path = os.path.join(root, f)
                        found_historical = True
                        break
                if found_historical:
                    break

    # 兜底：如果仍未设置 log_path 或文件不存在，则回退到实时日志
    if not log_path or not os.path.exists(log_path):
        log_path = real_time_log

    if not os.path.exists(log_path):
        return {
            "determined_stage": "L1",
            "suggested_find_commit_path": "oss-fuzz/projects/",
            "root_cause_analysis": f"Physical build log file is missing: {log_path}",
            "confidence_score": 0.5,
            "remediation_strategy": "Verify build log file generation."
        }

    # 读取日志并剔除 1+2+6 验证审计总结干扰
    try:
        with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
            raw_log = f.read()
    except Exception as e:
        return {
            "determined_stage": "L1",
            "suggested_find_commit_path": "oss-fuzz/projects/",
            "root_cause_analysis": f"Failed to read build log: {e}",
            "confidence_score": 0.5,
            "remediation_strategy": "Verify file access permissions."
        }

    val_marker = "--- VALIDATION SUMMARY"
    raw_compile_log = raw_log.split(val_marker)[0] if val_marker in raw_log else raw_log
    log_lines = [line.strip() for line in raw_compile_log.splitlines() if line.strip()]
    tail_500 = "\n".join(log_lines[-500:])

    # 1. 核心架构：反向级联搜索（L6 -> L5 -> L1 -> L2 -> L4 -> L3）
    def run_reverse_cascade_match(text: str) -> Optional[Tuple[str, float, str]]:
        # L6: Runtime 运行时自检层 (最高优先级)
        l6_sentinels = ["bad_build_check", "INFO: Seed:", "Loaded", "modules", "Atheris:", "Jazzer:"]
        l6_crashes = ["fuzz target exited", "AddressSanitizer:", "MemorySanitizer:", "TypeError:", "NameError:",
                      "java.lang.NoClassDefFoundError"]
        if any(sentinel in text for sentinel in l6_sentinels):
            if any(crash in text for crash in l6_crashes) or "exit status" in text:
                return "L6", 0.98, "Binary compiled successfully but crashed during runtime smoke check."

        # L5: Validation 构建产物交付规范层 (必须在成功哨兵后触发)
        l5_sentinels = ["BUILD SUCCESS", "Finished release", "Compiling .* done.", "ar .* done."]
        if any(re.search(sent, text, re.IGNORECASE) for sent in l5_sentinels):
            if any(err in text for err in
                   ["cp: cannot stat", "mv: target is not a directory", "chmod: cannot access", "zip I/O error"]):
                return "L5", 0.96, "Fuzzers compiled successfully, but copying or packaging into /out/ failed."
            if "out/" in text and "not found" in text:
                return "L5", 0.95, "Missing required packaging artifact or target directory."

        # L1: Bootstrap 物理环境层 (排除后续克隆成功以防误判)
        l1_patterns = [
            r"FROM\s+gcr.io/oss-fuzz-base/base-builder",
            r"E:\s*Unable\s+to\s+locate\s+package",
            r"add-apt-repository:\s*command\s+not\s+found",
            r"ERROR:\s*Service\s+'.*'\s+failed\s+to\s+build"
        ]
        if any(re.search(pat, text, re.IGNORECASE) for pat in l1_patterns):
            if "git clone" not in text:  # Negative Lookahead: 排除已正常 clone
                return "L1", 0.97, "Container environment bootstrap or apt dependencies installation failed."

        # L2: Dependency 依赖解析与构建树层
        l2_patterns = [
            r"git\s+clone.*fatal:",
            r"submodule.*failed",
            r"go\s+mod\s+download.*error",
            r"Updating\s+crates\.io.*failed",
            r"No\s+matching\s+distribution\s+found",
            r"pip\s+install.*failed",
            r"Downloading\s+from\s+central.*failed"
        ]
        if any(re.search(pat, text, re.IGNORECASE) for pat in l2_patterns):
            return "L2", 0.96, "Failed to clone upstream source repositories or fetch essential package dependencies."

        # L4: Linkage 驱动链接与打包层
        l4_patterns = [
            r"ld:\s+error:",
            r"undefined\s+reference\s+to",
            r"relocation\s+truncated",
            r"cannot\s+find\s+-l",
            r"libFuzzingEngine\.a\s+error",
            r"Missing\s+libFuzzer\s+main\s+symbol",
            r"pyinstaller.*error",
            r"overlapping\s+classes"
        ]
        if any(re.search(pat, text, re.IGNORECASE) for pat in l4_patterns):
            if "-c " not in text:  # Negative Lookahead: 排除单文件 -c 编译拼写错误
                return "L4", 0.98, "Symbols linkage failed. Undefined functions or missing static libraries."

        # L3: Compilation 静态插桩编译层 (最低优先级)
        l3_patterns = [
            r"clang\s+-c.*error:",
            r"fatal\s+error:",
            r"file\s+not\s+found",
            r"no\s+such\s+file\s+or\s+directory",
            r"cannot\s+open\s+include\s+file",
            r"missing\s+header",
            r"Compiling\s+.*error\s+\[E\d+\]:",
            r"javac.*cannot\s+find\s+symbol",
            r"cython.*error",
            r"syntax\s+error",
            r"expected\s+'.*'\s+before\s+",
            r"undeclared\s+identifier"
        ]
        if any(re.search(pat, text, re.IGNORECASE) for pat in l3_patterns):
            if "-o " in text and ("undefined reference" in text or "ld:" in text):
                return "L4", 0.90, "Downgraded to L4 Linkage warning."
            return "L3", 0.98, "Compiler syntax error, undeclared variables, or missing header files."

        return None

    # 第一阶段：尝试静态正向特征级联拦截
    static_decision = run_reverse_cascade_match(tail_500)
    if static_decision:
        stage, conf, analysis = static_decision
        # 工作区自动映射规则：Rule A 与 Rule B
        find_path = "process/project/" if stage in ["L3", "L6"] else "oss-fuzz/projects/"
        return {
            "determined_stage": stage,
            "suggested_find_commit_path": find_path,
            "root_cause_analysis": analysis,
            "confidence_score": conf,
            "remediation_strategy": f"Direct repair actions toward {stage} workspace."
        }

    # 第二阶段：收集三高价值信号
    sentinels = []
    for line in reversed(log_lines):
        if line.startswith("+ ") and not any(k in line.lower() for k in ["error", "fail", "exit status", "non-zero"]):
            sentinels.append(line)
            if len(sentinels) == 5:
                break
    sentinels.reverse()

    failed_cmd = ""
    for idx, line in enumerate(reversed(log_lines)):
        if "exit status" in line or "non-zero" in line or "command failed" in line.lower():
            start_pos = len(log_lines) - 1 - idx
            for u_idx in range(start_pos, -1, -1):
                if log_lines[u_idx].startswith("+ "):
                    failed_cmd = log_lines[u_idx]
                    break
            break

    whitelist_envs = ["SRC", "OUT", "WORK", "PROJECT_NAME", "ENGINE", "SANITIZER", "ARCHITECTURE", "CFLAGS", "CXXFLAGS",
                      "GOPATH", "PYTHONPATH", "JAVA_HOME"]
    env_vars = {k: os.environ.get(k, "N/A") for k in whitelist_envs}

    # 第三阶段：进入嵌套 LLM 仲裁
    if not dpseek_key:
        dpseek_key = os.getenv("DPSEEK_API_KEY")
    if not dpseek_key:
        return {"determined_stage": "L3", "confidence_score": 0.5,
                "root_cause_analysis": "Missing API key for LLM arbitration."}

    arbitration_prompt = f"""You are the world-class Cascade Build Stage Classifier (CBSC) Expert for the OSS-Fuzz automated repair platform. 
We have encountered an ambiguous or silent build failure. Analyze the highly-distilled context packet of high-value signals below:
[CONTEXT_START]
ENV_VARS: {json.dumps(env_vars, indent=2)}
SUCCESS_SEQUENCE: {json.dumps(sentinels, indent=2)}
FAILED_COMMAND: {failed_cmd}
ERROR_STACK: {tail_500[-4000:]}
[CONTEXT_END]
Analyze the context, map it to L1-L6 stages, and make the correct workspace routing decision.
Output exactly a single JSON. Do NOT include markdown wrappers outside the JSON block.
```json
{{
  "determined_stage": "L1/L2/L3/L4/L5/L6",
  "suggested_find_commit_path": "oss-fuzz/projects/<project_name>/ OR process/project/<project_name>/",
  "root_cause_analysis": "A concise 2-sentence description of the true failure cause and its path dependencies.",
  "confidence_score": 0.95,
  "remediation_strategy": "A brief actionable step for the solver to fix the issue."
}}
```"""

    try:
        response = litellm.completion(
            model="deepseek/deepseek-chat",
            messages=[{"role": "user", "content": arbitration_prompt}],
            temperature=0.2,
            api_key=dpseek_key
        )
        content = response.choices[0].message.content
        json_match = re.search(r'(\{[\s\S]*\})', content)
        if json_match:
            return json.loads(json_match.group(1))
        return json.loads(content.strip())
    except Exception as e:
        logger.error(f"LLM CBSC arbitration failed: {e}")
        return {
            "determined_stage": "L3",
            "suggested_find_commit_path": "process/project/",
            "root_cause_analysis": f"Arbitration errored: {e}",
            "confidence_score": 0.5,
            "remediation_strategy": "Fallback compile debug."
        }


def execute_hsr_decision(tool_context: ToolContext) -> dict:
    """
    Evaluates the Stage-Guided Decision Policy (HSR Engine).
    Compares SA >= SB dominance on the state tuple S = <L, V> to decide rollback action.
    Synchronizes double-workspace Git repositories with exact physical SHA targets on Rollback.
    """
    import subprocess
    import shutil
    import logging

    logger = logging.getLogger(__name__)
    session = tool_context.session
    ledger = TraceLedgerManager.load_ledger()
    print(f"[DEBUG HSR raw basic_information] {session.state.get('basic_information')}")
    basic_info = extract_basic_information(session.state.get("basic_information") or _LATEST_BASIC_INFORMATION)
    project_name = basic_info.get("project_name") or session.state.get("project_name") or session.state.get("project") or ledger.get("project_name") or "UNKNOWN"
    safe_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()

    default_source_path = os.path.join(os.getcwd(), "process", "project", safe_name) if safe_name else None
    default_config_path = os.path.join(os.getcwd(), "oss-fuzz", "projects", safe_name) if safe_name else None
    default_config_repo_path = os.path.join(os.getcwd(), "oss-fuzz") if safe_name else None

    project_source_path = basic_info.get("project_source_path") or session.state.get("project_source_path") or default_source_path
    project_config_path = basic_info.get("project_config_path") or session.state.get("project_config_path") or default_config_path
    project_config_repo_path = basic_info.get("project_config_repo_path") or session.state.get("project_config_repo_path") or default_config_repo_path

    project_source_path, project_config_path, project_config_repo_path = _coerce_project_paths_with_basic_information(
        project_source_path,
        project_config_path,
        basic_info
    )

    if project_source_path and not os.path.exists(project_source_path) and default_source_path and os.path.exists(default_source_path):
        project_source_path = default_source_path
    if project_config_path and not os.path.exists(project_config_path) and default_config_path and os.path.exists(default_config_path):
        project_config_path = default_config_path
    if project_config_repo_path and not os.path.exists(project_config_repo_path) and default_config_repo_path and os.path.exists(default_config_repo_path):
        project_config_repo_path = default_config_repo_path

    if project_config_path and not project_config_repo_path:
        candidate_root = os.path.abspath(os.path.join(project_config_path, "..", ".."))
        if os.path.exists(os.path.join(candidate_root, ".git")):
            project_config_repo_path = candidate_root

    session.state["project_name"] = project_name
    if project_source_path:
        session.state["project_source_path"] = project_source_path
    if project_config_path:
        session.state["project_config_path"] = project_config_path
    if project_config_repo_path:
        session.state["project_config_repo_path"] = project_config_repo_path
    print(f"[DBG] HSR context: project_name={project_name}, project_source_path={project_source_path}, project_config_path={project_config_path}, project_config_repo_path={project_config_repo_path}")

    if not ENABLE_ROLLBACK:
        nodes = ledger.get("nodes", [])
        pending_node_id = session.state.get("current_node_id")
        if pending_node_id is not None and pending_node_id > 0:
            pending_node = TraceLedgerManager.get_node_by_id(ledger, pending_node_id)
            parent_node = None
            if pending_node:
                parent_node = TraceLedgerManager.get_node_by_id(ledger, pending_node.get("parent_id", -1))
            if parent_node:
                parent_stage_after = parent_node.get("metrics", {}).get("build_stage_after", "N/A")
                parent_validation_after = parent_node.get("validation", {}).get("validation_report_after", {})
                TraceLedgerManager.update_node_fields(pending_node_id, {
                    "parent_id": parent_node["node_id"],
                    "metrics.build_stage_before": parent_stage_after if parent_stage_after not in [None, ""] else "N/A",
                    "validation.validation_report_before": parent_validation_after if isinstance(parent_validation_after, dict) else {}
                })
        return {"status": "success", "action": "NONE", "message": "HSR ablated. Continuing without rollback."}

    nodes = ledger.get("nodes", [])

    # 🔑 哨兵过滤：只选取已经回填了构建结论的活跃节点
    active_nodes = [n for n in nodes if n.get("metrics", {}).get("build_stage_after") not in [None, "N/A"]]

    if len(active_nodes) < 1:
        return {"status": "success", "action": "NONE",
                "message": "System at initial baseline. Build classification pending."}

    def get_validation_score(report: dict) -> int:
        score = 0
        if "pass" in str(report.get("step_1_official_list", "")).lower(): score += 1
        if "pass" in str(report.get("step_2_infra_compliance", "")).lower(): score += 1
        if "pass" in str(report.get("step_6_runtime_stability", "")).lower(): score += 1
        return score

    def dominates(L_a, V_a, L_b, V_b):
        return L_a > L_b or (L_a == L_b and V_a >= V_b)

    def _sync_pending_node_parent(parent_node: dict):
        pending_node_id = session.state.get("current_node_id")
        if pending_node_id is None or not parent_node:
            print(f"[DBG] HSR skip sync: pending_node_id={pending_node_id}, parent_node_exists={bool(parent_node)}")
            return

        parent_stage_after = parent_node.get("metrics", {}).get("build_stage_after", "N/A")
        parent_validation_after = parent_node.get("validation", {}).get("validation_report_after", {})

        print(f"[DBG] HSR writing parent fields: pending_node_id={pending_node_id}, parent_node_id={parent_node['node_id']}, parent_stage_after={parent_stage_after}, validation_before_keys={list(parent_validation_after.keys()) if isinstance(parent_validation_after, dict) else 'NON_DICT'}")

        TraceLedgerManager.update_node_fields(pending_node_id, {
            "parent_id": parent_node["node_id"],
            "metrics.build_stage_before": parent_stage_after if parent_stage_after not in [None, ""] else "N/A",
            "validation.validation_report_before": parent_validation_after if isinstance(parent_validation_after,
                                                                                         dict) else {}
        })

    curr_node = active_nodes[-1]
    print(f"[DBG] HSR curr_node_id={curr_node['node_id']}, curr_parent_id={curr_node.get('parent_id')}, session_current_node_id={session.state.get('current_node_id')}, active_node_ids={[n.get('node_id') for n in active_nodes]}")

    # Node 0 基准无需回滚，但仍需为下一悬空节点补父链
    if curr_node["node_id"] == 0:
        pending_node_id = session.state.get("current_node_id")
        if pending_node_id is not None and pending_node_id > curr_node["node_id"]:
            print(f"[DBG] HSR baseline sync for pending_node_id={pending_node_id} using parent_node_id=0")
            _sync_pending_node_parent(curr_node)
        return {"status": "success", "action": "NONE", "message": "Evaluating Node 0 baseline."}

    # 寻找父节点；若当前节点尚未写入 parent_id，则按 node_id-1 推导
    prev_node = next((n for n in nodes if n["node_id"] == curr_node.get("parent_id")), None)
    if not prev_node:
        inferred_parent_id = curr_node["node_id"] - 1
        prev_node = next((n for n in nodes if n["node_id"] == inferred_parent_id), None)
        print(f"[DBG] HSR prev_node missing for curr_node_id={curr_node['node_id']}, parent_id={curr_node.get('parent_id')}, inferred_parent_id={inferred_parent_id}, inferred_found={bool(prev_node)}")
        if not prev_node:
            return {"status": "success", "action": "NONE", "message": f"Node {curr_node['node_id']} has no stable parent."}

    # 判定逻辑
    stage_map = {"L1": 1, "L2": 2, "L3": 3, "L4": 4, "L5": 5, "L6": 6, "N/A": 0}
    L_curr = stage_map.get(curr_node.get("metrics", {}).get("build_stage_after", "N/A"), 0)
    L_prev = stage_map.get(prev_node.get("metrics", {}).get("build_stage_after", "N/A"), 0)

    V_curr = get_validation_score(curr_node.get("validation", {}).get("validation_report_after", {}))
    V_prev = get_validation_score(prev_node.get("validation", {}).get("validation_report_after", {}))

    is_curr_dominated = dominates(L_prev, V_prev, L_curr, V_curr)

    decision_status = "Stable"
    rollback_type = "NONE"
    should_rollback = False
    target_node = None

    if is_curr_dominated:
        should_rollback = True
        if L_curr == L_prev and V_curr == V_prev:
            rollback_type = "SINGLE_STEP"
            # 规则：如果 Node i 被标记为 SINGLE_STEP 回退，则下一待处理节点挂回 Node i 的父节点。
            target_node = prev_node
            decision_status = "Neutral Path"
        else:
            rollback_type = "ADAPTIVE"
            decision_status = "Degrading"
            # 规则：ADAPTIVE 回退时，向前寻找一个状态严格优于 Node i 的更早节点。
            historical_candidates = [n for n in nodes if n.get("node_id", -1) < curr_node["node_id"]]
            for node in reversed(historical_candidates):
                L_hist = stage_map.get(node.get("metrics", {}).get("build_stage_after", "N/A"), 0)
                V_hist = get_validation_score(node.get("validation", {}).get("validation_report_after", {}))
                is_strictly_better = L_hist > L_curr or (L_hist == L_curr and V_hist > V_curr)
                if is_strictly_better:
                    target_node = node
                    break
            if target_node is None: target_node = nodes[0]

    TraceLedgerManager.update_node_fields(curr_node["node_id"], {
        "identification.node_status": decision_status,
        "identification.should_rollback": should_rollback,
        "identification.rollback_type": rollback_type
    })
    print(f"[DBG] HSR decision: curr_node_id={curr_node['node_id']}, should_rollback={should_rollback}, rollback_type={rollback_type}, target_node_id={target_node['node_id'] if target_node else None}")

    if should_rollback and target_node:
        oss_sha = target_node["git_sha_state"].get("oss-fuzz_sha")
        prj_sha = target_node["git_sha_state"].get("project_sha")
        print(f"[DBG] HSR rollback target: target_node_id={target_node['node_id']}, oss_sha={oss_sha}, prj_sha={prj_sha}")

        # 🔑 容错逻辑：若 SHA 无效则跳过物理回滚
        if not oss_sha or oss_sha == "N/A" or not prj_sha or prj_sha == "N/A":
            print(f"--- [HSR Warning] SHA invalid (OSS:{oss_sha}, PRJ:{prj_sha}), skipping physical reset ---")
            session.state["rollback_triggered"] = True
            _sync_pending_node_parent(target_node)
            return {"status": "success", "action": "ROLLBACK_LOGICAL_ONLY", "message": "Logical rollback only."}


        # 执行物理回滚
        shas_to_reset = [(project_source_path, prj_sha), (project_config_repo_path, oss_sha)]
        for repo_path, target_sha in shas_to_reset:
            print(f"[DBG] HSR physical reset candidate: repo_path={repo_path}, target_sha={target_sha}")
            if not repo_path:
                print(f"--- [HSR Warning] Missing repo_path for target_sha={target_sha}. Skipping this reset target. ---")
                continue
            if os.path.exists(repo_path):
                # 🔑 联动避险：在执行 git clean 之前，先主动夺回整个工作区的权限，防止 git 报 Permission Denied
                reclaim_path_permissions(repo_path)

                subprocess.run(["git", "-C", repo_path, "reset", "--hard", target_sha], check=True, capture_output=True)
                subprocess.run(["git", "-C", repo_path, "clean", "-fxd"], check=True, capture_output=True)
            else:
                print(f"--- [HSR Warning] repo_path does not exist on disk: {repo_path}. Skipping this reset target. ---")

        if project_config_path:
            out_dir = os.path.join(project_config_path, "..", "..", "build", "out", project_name)
            safe_delete_path(out_dir)
        else:
            print("--- [HSR Warning] project_config_path missing; skipping build/out cleanup. ---")

        clear_commit_analysis_state()
        session.state["rollback_triggered"] = True
        session.state["software_sha"] = prj_sha
        session.state["oss_fuzz_sha"] = oss_sha

        ledger_after_reset = TraceLedgerManager.load_ledger()
        remaining_ids = [n.get("node_id") for n in ledger_after_reset.get("nodes", [])]
        print(f"[DBG] HSR full ledger preserved with node_ids={remaining_ids}")

        _sync_pending_node_parent(target_node)
        return {"status": "success", "action": "ROLLBACK", "target_node_id": target_node["node_id"],
                "message": "Physical rollback successful."}

        # NONE 路径
    _sync_pending_node_parent(curr_node)
    return {"status": "success", "action": "NONE", "message": "System stable."}


def clear_commit_analysis_state() -> Dict[str, str]:
    """
    Remove the commit analysis sentinel file to allow commit_finder_agent to re-run in the next loop.
    """
    commit_analysis_file = "generated_prompt_file/commit_changed.txt"
    if os.path.exists(commit_analysis_file):
        try:
            safe_delete_path(commit_analysis_file)
            return {"status": "success",
                    "message": f"Cleared old commit analysis state. '{commit_analysis_file}' has been removed."}
        except Exception as e:
            return {"status": "error", "message": f"Failed to remove '{commit_analysis_file}': {e}"}
    else:
        return {"status": "success", "message": "No commit analysis state to clear."}


# rsmc_agent tools

def init_or_update_rsmc_ledger(tool_context: ToolContext, solved_problems: str, unsolved_problems: str,
                               reflection_analysis: str, loop_summary: str) -> dict:
    session = tool_context.session
    current_node_id = session.state.get("current_node_id", 0)

    ledger = TraceLedgerManager.load_ledger()

    if not ENABLE_REFLECTION:
        last_report = session.state.get("last_validation_report", {})
        bitmap_keys = [
            "step_1_official_list",
            "step_2_infra_compliance",
            "step_3_sanitizer_injected",
            "step_4_engine_control",
            "step_5_logic_linkage",
            "step_6_runtime_stability"
        ]
        step_1_6_bitmap = [
            1 if str(last_report.get(k, "")).startswith("pass") else 0
            for k in bitmap_keys
        ]

        TraceLedgerManager.update_node_fields(current_node_id, {
            "action_and_intent.loop_summary": "RSMC ablated. Semantic reflection skipped.",
            "validation.validation_report_after": last_report,
            "validation.step_1_6_bitmap": step_1_6_bitmap,
            "semantic_memory.solved_problems": "RSMC ablated.",
            "semantic_memory.unsolved_problems": "RSMC ablated.",
            "semantic_memory.reflection_analysis": "RSMC ablated by configuration."
        })

        pending_node = next((
            n for n in ledger.get("nodes", [])
            if n.get("parent_id") == current_node_id and n.get("metrics", {}).get("build_stage_after") is None
        ), None)
        if pending_node is None:
            latest_ledger = TraceLedgerManager.load_ledger()
            next_node_id, latest_ledger = TraceLedgerManager.allocate_next_node_id(latest_ledger)
            new_node = {
                "node_id": next_node_id,
                "parent_id": current_node_id,
                "identification": {
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "node_status": "Stable",
                    "should_rollback": False,
                    "rollback_type": "NONE"
                },
                "git_sha_state": {"oss-fuzz_sha": None, "project_sha": None},
                "action_and_intent": {
                    "root_cause_commit_sha": "N/A",
                    "active_workspace": "UNKNOWN",
                    "target_file": "N/A",
                    "repair_strategy": "Pending",
                    "loop_summary": "Pending"
                },
                "metrics": {
                    "Ldel": 0, "Ladd": 0,
                    "build_stage_before": "N/A",
                    "build_stage_after": None
                },
                "validation": {
                    "step_1_6_bitmap": None,
                    "validation_report_before": {},
                    "validation_report_after": {}
                },
                "semantic_memory": {
                    "solved_problems": None,
                    "unsolved_problems": None,
                    "reflection_analysis": None
                }
            }
            latest_ledger["nodes"].append(new_node)
            TraceLedgerManager.save_ledger(latest_ledger)
        else:
            next_node_id = pending_node.get("node_id")
        session.state["current_node_id"] = next_node_id
        return {"status": "success", "message": f"RSMC ablated. Node {current_node_id} kept minimal ledger integrity, Node {next_node_id} ready."}

    # Prefer the in-memory validation snapshot; if the session view is stale/empty,
    # preserve the report that was already backfilled by the orchestrator.
    ledger_current_node = next((n for n in ledger.get("nodes", []) if n.get("node_id") == current_node_id), {})
    existing_validation_after = ledger_current_node.get("validation", {}).get("validation_report_after", {})
    last_report = session.state.get("last_validation_report", {})
    if not last_report and existing_validation_after:
        last_report = existing_validation_after

    # 🔑 从 last_validation_report 计算 step_1_6_bitmap
    bitmap_keys = [
        "step_1_official_list",
        "step_2_infra_compliance",
        "step_3_sanitizer_injected",
        "step_4_engine_control",
        "step_5_logic_linkage",
        "step_6_runtime_stability"
    ]
    step_1_6_bitmap = [
        1 if str(last_report.get(k, "")).startswith("pass") else 0
        for k in bitmap_keys
    ]

    filled_fields = {
        "action_and_intent.loop_summary": loop_summary[:800].strip(),
        "validation.validation_report_after": last_report,
        "validation.step_1_6_bitmap": step_1_6_bitmap,
        "semantic_memory.solved_problems": solved_problems[:150].strip(),
        "semantic_memory.unsolved_problems": unsolved_problems[:150].strip(),
        "semantic_memory.reflection_analysis": reflection_analysis[:800].strip()
    }
    TraceLedgerManager.update_node_fields(current_node_id, filled_fields)

    # 幂等性开辟新节点
    pending_node = next((
        n for n in ledger.get("nodes", [])
        if n.get("parent_id") == current_node_id and n.get("metrics", {}).get("build_stage_after") is None
    ), None)
    next_node_id = pending_node.get("node_id") if pending_node else None
    print(f"[DBG] RSMC current_node_id={current_node_id}, next_node_id={next_node_id}")
    if pending_node is None:
        latest_ledger = TraceLedgerManager.load_ledger()
        next_node_id, latest_ledger = TraceLedgerManager.allocate_next_node_id(latest_ledger)
        new_node = {
            "node_id": next_node_id,
            "parent_id": current_node_id,
            "identification": {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "node_status": "Stable",
                "should_rollback": False,
                "rollback_type": "NONE"
            },
            "git_sha_state": {"oss-fuzz_sha": None, "project_sha": None},
            "action_and_intent": {
                "root_cause_commit_sha": "N/A",
                "active_workspace": "UNKNOWN",
                "target_file": "N/A",
                "repair_strategy": "Pending",
                "loop_summary": "Pending"
            },
            "metrics": {
                "Ldel": 0, "Ladd": 0,
                "build_stage_before": "N/A",
                "build_stage_after": None
            },
            "validation": {
                "step_1_6_bitmap": None,
                "validation_report_before": {},
                "validation_report_after": {}
            },
            "semantic_memory": {
                "solved_problems": None,
                "unsolved_problems": None,
                "reflection_analysis": None
            }
        }

        latest_ledger["nodes"].append(new_node)
        TraceLedgerManager.save_ledger(latest_ledger)
        print(f"[DBG] RSMC created node_id={new_node['node_id']}, parent_id={new_node['parent_id']}, build_stage_before={new_node['metrics']['build_stage_before']}, validation_report_before={new_node['validation']['validation_report_before']}")

    session.state["current_node_id"] = next_node_id
    print(f"[DBG] RSMC session current_node_id updated to {session.state['current_node_id']}")
    return {"status": "success", "message": f"Node {current_node_id} backfilled, Node {next_node_id} ready."}


def query_trace_ledger(tool_context: ToolContext, field_keys: List[str], node_id: Optional[int] = None) -> dict:
    """
    Secure field-level getter tool for both LLM Agents and python backend prompter.
    Supports dot-notation path parsing and strictly accesses data based on the node_id.
    """
    session = tool_context.session
    ledger = TraceLedgerManager.load_ledger()

    if node_id is None:
        session_node_id = session.state.get("current_node_id", 0)
        session_node = TraceLedgerManager.get_node_by_id(ledger, session_node_id)
        if session_node:
            node_id = session_node.get("parent_id", -1)
        else:
            node_id = -1

    # 🔑 修改：账本不存在或 nodes 为空时，返回空数据而非 error
    if not ledger or not ledger.get("nodes"):
        retrieved_data = {key: "N/A (Ledger not initialized)" for key in field_keys}
        return {
            "status": "success",
            "node_id": node_id,
            "data": retrieved_data
        }

    target_node = TraceLedgerManager.get_node_by_id(ledger, node_id)

    # 🔑 修改：node_id 不存在时，返回空数据而非 error
    if not target_node:
        retrieved_data = {key: "N/A (Node not yet created)" for key in field_keys}
        return {
            "status": "success",
            "node_id": node_id,
            "data": retrieved_data
        }

    retrieved_data = {}
    for key_path in field_keys:
        parts = key_path.split('.')
        current_level = target_node
        success = True
        for part in parts:
            if isinstance(current_level, dict) and part in current_level:
                current_level = current_level[part]
            else:
                success = False
                break
        if success:
            retrieved_data[key_path] = current_level
        else:
            retrieved_data[key_path] = "N/A (Field not instantiated)"

    return {
        "status": "success",
        "node_id": node_id,
        "data": retrieved_data
    }




def extract_buggy_line_info(log_path: str, project_name: str = "", project_source_path: str = "",
                            error_date: str = "") -> dict:
    """
    [HAFix Phase 1 & 2] Dynamic Clue Mining + Self-Healing Identification.
    Replaces the original logic to support multi-mode extraction and path scoring.

    Args:
        log_path: Path to the build log.
        project_name: Name of the project.
        project_source_path: (New) Root path of the source code for path validation.
        error_date: (New) Date string for fallback time-window search.
    """
    import os, re, subprocess
    if not os.path.exists(log_path): return {"status": "error", "message": "Log file not found."}

    # --- Helper: Read tail lines with noise filtering ---
    def read_log_tail(path, count):
        try:
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            # Filter out Agent diagnostic noise and common non-error logs
            noise = ["--- Tool:", "RESULT:", "[⚠️", "usage: helper.py", "Step #"]
            return [l for l in lines[-count:] if not any(kw in l for kw in noise)]
        except:
            return []

    # --- Phase 1: Clue Mining ---
    # 1. Scan 500 lines
    content = "".join(read_log_tail(log_path, 500))
    # Regex supports C/C++/Go/Rust/Config files: (path/file.ext:line:)
    pattern = r"([\w\-\./_]+\.(?:c|cpp|h|cc|cxx|rs|go|py|sh|java|swift|cmake|txt)):(\d+):?"
    matches = re.findall(pattern, content)

    # 2. Fallback to 1000 lines if empty
    if not matches:
        content = "".join(read_log_tail(log_path, 1000))
        matches = re.findall(pattern, content)

    # 3. Keyword Fallback for Phase 2 Path B
    if not matches:
        keywords = [kw for kw in ["GOMODCACHE", "WORKDIR", "overlay", "lib.*not found", "undefined reference"] if
                    re.search(kw, content, re.I)]
        return {"status": "success", "clue_type": "keyword", "data": {"keywords": keywords, "error_date": error_date}}

    # --- Phase 2: Path Self-Healing & Scoring ---
    # Pre-load recent changes for scoring (+5 points)
    recent_changes = set()
    if project_source_path and os.path.isdir(os.path.join(project_source_path, ".git")):
        try:
            res = subprocess.run(["git", "-C", project_source_path, "log", "-n", "50", "--name-only", "--format="],
                                 capture_output=True, text=True, timeout=10)
            recent_changes = {f.strip() for f in res.stdout.split('\n') if f.strip()}
        except:
            pass

    scored_candidates = []
    for raw_file, raw_line in matches:
        score, final_path = 0, raw_file

        # Check direct existence (Score 100)
        if project_source_path and os.path.exists(os.path.join(project_source_path, raw_file)):
            score = 100
        else:
            # Attempt to find file via search (Self-Healing)
            basename = os.path.basename(raw_file)
            if project_source_path:
                try:
                    find_cmd = ["find", project_source_path, "-name", basename, "-type", "f"]
                    find_res = subprocess.run(find_cmd, capture_output=True, text=True, timeout=5).stdout.strip().split(
                        '\n')
                    best_s, best_c = -999, None
                    for cand in [c for c in find_res if c]:
                        rel = os.path.relpath(cand, project_source_path)
                        s = 0
                        if os.path.dirname(raw_file) in rel: s += 10  # +10: Parent dir match
                        if rel in recent_changes: s += 5  # +5: Recently modified
                        s -= abs(rel.count('/') - raw_file.count('/'))  # -1: Depth penalty
                        if s > best_s: best_s, best_c = s, rel
                    if best_c: score, final_path = 60 + best_s, best_c
                except:
                    pass

        scored_candidates.append({"file": final_path, "line": int(raw_line), "score": score})

    scored_candidates.sort(key=lambda x: x['score'], reverse=True)
    best = scored_candidates[0] if scored_candidates else None

    # Execute Blame (if score >= 60)
    if best and best['score'] >= 60 and project_source_path:
        try:
            blame_cmd = ["git", "-C", project_source_path, "blame", "-L", f"{best['line']},{best['line']}",
                         "--porcelain", best['file']]
            res = subprocess.run(blame_cmd, capture_output=True, text=True, check=True, timeout=10)
            sha = res.stdout.split('\n')[0].split(' ')[0]
            if len(sha) >= 7:
                return {"status": "success", "clue_type": "blame",
                        "data": {"sha": sha, "file": best['file'], "line": best['line']}}
        except:
            pass

    # Fallback to Time-Window Suspects
    return {"status": "success", "clue_type": "time_window",
            "data": {"file": best['file'] if best else None, "error_date": error_date}}


def get_enhanced_history_context(project_source_path: str, clue_data: dict = None, file_rel_path: str = "",
                                 line_num: int = 0, sha: str = "") -> dict:
    """
    [HAFix Phase 3] Chain-of-Evidence Synthesis.
    Replaces the original logic to support multi-mode evidence gathering.

    Args:
        project_source_path: Root path of the source code.
        clue_data: (New) Structured output from Phase 1 (extract_buggy_line_info).
        file_rel_path: (Legacy/Deprecated) Used if clue_data is missing.
        line_num: (Legacy/Deprecated) Used if clue_data is missing.
        sha: (Legacy/Deprecated) Used if clue_data is missing.
    """
    import os, subprocess
    from datetime import datetime, timedelta

    if not os.path.isdir(os.path.join(project_source_path, ".git")):
        return {"status": "error", "message": "Not a git repository."}

    # --- Auto-convert Legacy Call to Phase 1 Data if clue_data is missing ---
    if not clue_data:
        if sha:
            clue_data = {"clue_type": "blame", "data": {"sha": sha, "file": file_rel_path}}
        elif file_rel_path and line_num:
            clue_data = {"clue_type": "time_window", "data": {"file": file_rel_path}}  # Fallback handling

    if not clue_data:
        return {"status": "error", "message": "No clue data provided."}

    clue_type = clue_data.get("clue_type")
    payload = clue_data.get("data", {})
    evidence = {"clue_type": clue_type, "suspect_sha": payload.get("sha", "N/A"), "core_files": [],
                "auxiliary_timeline": [], "diffs": []}

    try:
        # 1. Determine Core Tracing Files
        if clue_type == "blame":
            target_sha = payload['sha']
            show_res = subprocess.run(
                ["git", "-C", project_source_path, "show", "--name-only", "--format=", target_sha],
                capture_output=True, text=True, timeout=10).stdout
            changed = [f.strip() for f in show_res.split('\n') if f.strip()]
            # Filter to top 3 relevant source/config files
            exts = ('.c', '.go', '.cpp', '.h', '.sh', 'Dockerfile', 'build.sh', 'go.mod', 'CMakeLists.txt')
            evidence["core_files"] = [f for f in changed if f.endswith(exts) or any(x in f for x in exts)][:3]
        else:
            # Keyword/Time-Window mode: prioritize the reported file
            if payload.get("file"):
                evidence["core_files"] = [payload.get("file")]
            # If no file, leave empty for Agent to scan config

        # 2. Build Time Window (±24h)
        error_date = payload.get("error_date", "")
        since_until = []
        if error_date and error_date.strip():
            try:
                clean_date = error_date.replace('.', '-').replace('/', '-')
                t = datetime.strptime(clean_date.split()[0], '%Y-%m-%d')
                since_until = [f"--since={(t - timedelta(days=1)).strftime('%Y-%m-%d')}",
                               f"--until={(t + timedelta(days=1)).strftime('%Y-%m-%d')}"]
            except:
                pass

        # 3. Chain-of-Evidence Collection
        for f in [x for x in evidence["core_files"] if x and os.path.exists(os.path.join(project_source_path, x))]:
            # A. Auxiliary Timeline (git log -n 5)
            log_cmd = ["git", "-C", project_source_path, "log", *since_until, "-n", "5", "--format=%H|%cd|%s", "--", f]
            log_res = subprocess.run(log_cmd, capture_output=True, text=True, timeout=10).stdout.strip()
            if log_res:
                evidence["auxiliary_timeline"].append(
                    {"file": f, "commits": [l.split('|') for l in log_res.split('\n') if '|' in l]})

            # B. Structural Sampling (Unified Diff -U3)
            if evidence["suspect_sha"] != "N/A" and len(evidence["suspect_sha"]) >= 7:
                diff_cmd = ["git", "-C", project_source_path, "show", "-U3", "--format=", evidence["suspect_sha"], "--",
                            f]
                diff_res = subprocess.run(diff_cmd, capture_output=True, text=True, timeout=10).stdout
                evidence["diffs"].append({"file": f, "content": diff_res[:8000]})  # 8000 char Token Guard

        return {"status": "success", "data": evidence}
    except Exception as e:
        return {"status": "error", "message": f"Synthesis failed: {str(e)}"}


def checkout_project_commit(project_source_path: str, sha: str) -> Dict[str, str]:
    """
    Executes git checkout with built-in remote self-healing.
    If the SHA is not found locally (reference is not a tree), it attempts to fetch from origin.
    """
    import os
    import subprocess

    print(f"--- Tool: checkout_project_commit | SHA: {sha} | Path: {project_source_path} ---")

    if not os.path.isdir(os.path.join(project_source_path, ".git")):
        return {'status': 'error', 'message': f"Path '{project_source_path}' is not a valid git repository."}

    try:
        # 1. 预清理：强制放弃任何本地残留修改，确保切换环境绝对干净
        subprocess.run(["git", "reset", "--hard", "HEAD"], capture_output=True, cwd=project_source_path)
        subprocess.run(["git", "clean", "-fdx"], capture_output=True, cwd=project_source_path)

        # 2. 尝试执行物理切换
        command = ["git", "checkout", sha]
        result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8', cwd=project_source_path)

        # 3. 🔑 核心补丁：处理“引用不是一个树”的错误 (Commit 不在本地)
        if result.returncode != 0:
            err_msg = result.stderr.lower()
            if "reference is not a tree" in err_msg or "not a commit" in err_msg or "引用不是一个树" in result.stderr:
                print(f"--- [SELF-HEALING] SHA {sha} missing. Attempting to fetch from remote... ---")

                # 尝试从远程精准拉取该 SHA 节点
                fetch_cmd = ["git", "fetch", "origin", sha]
                fetch_res = subprocess.run(fetch_cmd, capture_output=True, text=True, cwd=project_source_path)

                if fetch_res.returncode == 0:
                    print(f"--- [SELF-HEALING] Fetch successful. Retrying checkout... ---")
                    # 再次尝试 checkout
                    result = subprocess.run(["git", "checkout", sha], capture_output=True, text=True,
                                            cwd=project_source_path)
                else:
                    # 如果精准 fetch 失败，尝试全量 unshallow (针对部分浅克隆仓库)
                    print(f"--- [SELF-HEALING] Precise fetch failed. Attempting unshallow fetch... ---")
                    subprocess.run(["git", "fetch", "--unshallow"], capture_output=True, cwd=project_source_path)
                    result = subprocess.run(["git", "checkout", sha], capture_output=True, text=True,
                                            cwd=project_source_path)

        # 4. 最终结果判定
        if result.returncode == 0:
            # 记录成功后的当前状态摘要
            head_info = subprocess.run(["git", "rev-parse", "HEAD"], capture_output=True, text=True,
                                       cwd=project_source_path).stdout.strip()
            return {
                'status': 'success',
                'message': f"Successfully checked out SHA {sha}. Current HEAD: {head_info}"
            }
        else:
            return {
                'status': 'error',
                'message': f"Git checkout failed after self-healing attempts. Error: {result.stderr.strip()}"
            }

    except Exception as e:
        return {'status': 'error', 'message': f"Unexpected system error during checkout: {str(e)}"}


def download_remote_log(log_url: str, project_name: str, error_time_str: str) -> Dict[str, str]:
    """
    Download remote log file and save it locally using 'YYYY_M_D error.txt' format.
    🔑 优化：提供多维时间分隔符(.-/)自适应健壮解析；
    🔑 优化：缓存命中时实现零网络请求，即刻放行。
    """
    import os
    import sys
    import requests
    from datetime import datetime

    print(f"--- Tool: download_remote_log called for URL: {log_url} ---")

    try:
        # 🔑 1. 强鲁棒日期分隔符清洗自愈
        clean_time_str = error_time_str.strip().replace('.', '-').replace('/', '-')
        error_date = None

        for fmt in ['%Y-%m-%d', '%Y_%m_%d', '%Y%m%d']:
            try:
                error_date = datetime.strptime(clean_time_str, fmt).date()
                break
            except ValueError:
                continue

        if not error_date:
            return {'status': 'error',
                    'message': f"Failed to parse error date string '{error_time_str}' with standard formats."}

        # 确定本地缓存路径
        local_log_dir = os.path.join("build_error_log", project_name)
        os.makedirs(local_log_dir, exist_ok=True)

        if sys.platform == "win32":
            local_log_filename = error_date.strftime("%Y_%#m_%#d") + " error.txt"
        else:
            local_log_filename = error_date.strftime("%Y_%-m_%-d") + " error.txt"

        local_log_filepath = os.path.join(local_log_dir, local_log_filename)

        # 🔑 2. 缓存就地放行逻辑：如果已存在，零网络损耗
        if os.path.exists(local_log_filepath) and os.path.getsize(local_log_filepath) > 0:
            print(f"--- Log file already exists locally: {local_log_filepath}. Skipping download. ---")
            return {"status": "success", "local_path": os.path.abspath(local_log_filepath),
                    "message": "Local log cache hit. Skipping remote pull."}

        # 3. 缓存未命中，执行网络下载
        print(f"--- Downloading log from {log_url} to {local_log_filepath} ---")
        response = requests.get(log_url, stream=True, timeout=30)
        response.raise_for_status()

        with open(local_log_filepath, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)

        print(f"--- Successfully downloaded log to: {local_log_filepath} ---")
        return {"status": "success", "local_path": os.path.abspath(local_log_filepath),
                "message": "Successfully downloaded remote log."}

    except requests.exceptions.RequestException as e:
        return {"status": "error", "message": f"Failed to download log from {log_url}: {e}"}
    except Exception as e:
        return {"status": "error", "message": f"An unexpected error occurred during log download: {e}"}


def update_reflection_journal(
        project_name: str,
        attempt_id: int,
        round_id: int,
        strategy_used: str,
        solution_plan: str,
        build_log_tail: str,
        reflection_analysis: str,
        deterioration_score: int,
        solved_problems: str,
        unsolved_problems: str,
        should_rollback: bool = False
) -> Dict:
    """
    Explicitly record Attempt and Round IDs, store concise problem descriptions, and extract recent lessons for the state.
    """
    import os
    import json
    from datetime import datetime

    if not os.environ.get("ENABLE_REFLECTION", "True") == "True":
        return {"status": "success", "trigger_rollback": False}

    print(f"--- Tool: update_reflection_journal (v5) for A{attempt_id}_R{round_id} ---")
    JOURNAL_FILE = "reflection_journal.json"

    new_entry = {
        "attempt_id": attempt_id,
        "round_id": round_id,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "strategy": strategy_used,
        "solved": solved_problems,
        "unsolved": unsolved_problems,
        "deterioration_score": deterioration_score,
        "reflection": reflection_analysis,
        "should_rollback": should_rollback
    }

    history = []
    if os.path.exists(JOURNAL_FILE):
        try:
            with open(JOURNAL_FILE, 'r', encoding='utf-8') as f:
                history = json.load(f)
        except:
            pass
    history.append(new_entry)

    with open(JOURNAL_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

    current_attempt_history = [h for h in history if h['attempt_id'] == attempt_id]
    consecutive_high_score = False
    if len(current_attempt_history) >= 2:
        if current_attempt_history[-1].get("deterioration_score", 0) > 7 and \
                current_attempt_history[-2].get("deterioration_score", 0) > 7:
            consecutive_high_score = True

    lessons = []
    for h in current_attempt_history[-3:]:
        lessons.append(
            f"A{h['attempt_id']}_R{h['round_id']} (Score:{h['deterioration_score']}):\n"
            f"  [Fixed]: {h['solved']}\n"
            f"  [Pending]: {h['unsolved']}"
        )
    summary_for_state = "\n".join(lessons)

    return {
        "status": "success",
        "reflection_summary": summary_for_state,
        "trigger_rollback": should_rollback or consecutive_high_score,
        "deterioration_score": deterioration_score
    }



def few_shot_rag_retrieve(expert_knowledge_path: str, log_path: str) -> dict:
    """
    Three-Step Few-shot RAG Retrieval Pipeline.

    Step 1: Log Error Pattern Matching
      - Reads [FAILURE_REGION] from commit_changed.txt.
      - Falls back to tail of fuzz_build_log.txt if missing/empty.
      - Runs positive pattern and optional negative exclude_pattern regex matches.

    Step 2: Keywords Association & Guideline Expansion
      - Extracts keywords from matched ERR_xxx entries.
      - Intersects these keywords with keywords of expert_guidelines to fetch associated GL_xxx.

    Step 3: Prompt Injection & Severity Ordinal Sorting
      - Merges matched ERRs and associated GLs.
      - Associated GLs dynamically inherit the highest severity weight of the matched ERRs that triggered them.
      - Sorts by severity weight descending and formats Top 4 priority blocks for injection.
    """
    if not ENABLE_EXPERT_KNOWLEDGE:
        return {
            "status": "success",
            "message": "Few-shot RAG disabled by ablation.",
            "rag_context": "",
            "matched_errors_count": 0
        }

    if not os.path.exists(expert_knowledge_path):
        return {
            "status": "error",
            "message": f"Expert knowledge database missing at: {expert_knowledge_path}",
            "rag_context": ""
        }

    try:
        with open(expert_knowledge_path, 'r', encoding='utf-8') as f:
            kb = json.load(f)
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to load expert knowledge database: {e}",
            "rag_context": ""
        }

    # =================================================================
    # Step 1: 日志报错特征匹配 (Failure Region Scanning)
    # =================================================================
    failure_region = ""
    changed_file = "generated_prompt_file/commit_changed.txt"

    # 优先读取 ECRCL 生成的精确 FAILURE_REGION
    if os.path.exists(changed_file) and os.path.getsize(changed_file) > 0:
        try:
            with open(changed_file, 'r', encoding='utf-8') as f:
                content = f.read()
            # 🔑 修复：使用正确的非贪婪匹配与中括号安全转义
            match = re.search(r"\[FAILURE_REGION\]\s*([\s\S]*?)\s*\[ATTRIBUTION_TYPE\]", content)
            if match:
                failure_region = match.group(1).strip()
        except Exception as e:
            logger.debug(f"Failed to parse FAILURE_REGION from commit_changed.txt: {e}")

    # 状态自愈：若归因工件缺失，自适应切换到原始编译日志尾部切片
    if not failure_region and os.path.exists(log_path):
        try:
            with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
                log_lines = f.readlines()
            # 提取尾部 200 行编译日志切片作为特征源
            failure_region = "".join(log_lines[-200:])
        except Exception as e:
            logger.error(f"Failed to read fallback compilation log tail: {e}")

    if not failure_region:
        return {
            "status": "success",
            "message": "No compilation failure context available for RAG.",
            "rag_context": ""
        }

    # 执行正负向双重正则特征匹配
    matched_errors = []
    for err in kb.get("error_patterns", []):
        pattern = err.get("pattern")
        if not pattern:
            continue
        try:
            # 正向特征匹配 ("眼睛" 扫描)
            if re.search(pattern, failure_region, re.IGNORECASE):
                # 负向噪音过滤
                exclude_pattern = err.get("exclude_pattern")
                if exclude_pattern and re.search(exclude_pattern, failure_region, re.IGNORECASE):
                    # 命中了负向噪音特征，直接排除
                    continue
                matched_errors.append(err)
        except Exception as e:
            logger.error(f"Regex pattern match error for {err.get('id', 'UNKNOWN')}: {e}")

    # =================================================================
    # Step 2: 关键词关联扩展 (Keywords Intersection)
    # =================================================================
    matched_guidelines = []

    # 建立 ERR_id 与其 keywords 关联映射，便于后续 GL 继承严重度
    err_keyword_to_severity: Dict[str, str] = {}
    for err in matched_errors:
        severity = err.get("severity", "MINOR")
        for kw in err.get("keywords", []):
            kw_clean = kw.lower().strip()
            # 🔑 修复：修正错行问题，同一个关键词对应多个 ERR 时保留最高严重度
            if kw_clean not in err_keyword_to_severity:
                err_keyword_to_severity[kw_clean] = severity
            else:
                curr_sev = err_keyword_to_severity[kw_clean]
                weight_map = {"CRITICAL": 3, "MAJOR": 2, "MINOR": 1}
                if weight_map.get(severity, 0) > weight_map.get(curr_sev, 0):
                    err_keyword_to_severity[kw_clean] = severity

    # 匹配关联的 GL_xxx 条目并进行严重度隐式绑定
    for gl in kb.get("expert_guidelines", []):
        gl_kws = [kw.lower().strip() for kw in gl.get("keywords", [])]

        # 计算关键词交集
        intersected_kws = set(gl_kws).intersection(set(err_keyword_to_severity.keys()))
        if intersected_kws:
            # 拓扑继承设计：寻找关联的最高严重度并赋予该 GL
            max_severity = "MINOR"
            weight_map = {"CRITICAL": 3, "MAJOR": 2, "MINOR": 1}
            for kw in intersected_kws:
                sev_cand = err_keyword_to_severity[kw]
                if weight_map.get(sev_cand, 0) > weight_map.get(max_severity, 0):
                    max_severity = sev_cand

            # 浅拷贝复制一份，写入继承得到的伪严重度属性用于统一排序
            gl_copied = dict(gl)
            gl_copied["_inherited_severity"] = max_severity
            matched_guidelines.append(gl_copied)

    # =================================================================
    # Step 3: Prompt 注入与严重等级优先级排序 (Prompt Injection)
    # =================================================================
    # 硬编码严重级别优先级字典（序数权重换算）
    SEVERITY_WEIGHT = {
        "CRITICAL": 3,
        "MAJOR": 2,
        "MINOR": 1
    }

    # 统一整合包结构
    unified_candidates: List[Tuple[str, int, str]] = []

    # 1. 压入命中错误模式 (ERR)
    for err in matched_errors:
        weight = SEVERITY_WEIGHT.get(err.get("severity", "MINOR"), 1)
        block_text = f"""[MATCHED ERROR PATTERN: {err['id']} ({err['name']})]
Severity: {err['severity']}
Diagnosis: {err['diagnosis']}
Remediation Action: {err['remediation']['action']}
Remediation Rule: {err['remediation']['rule']}
Verification: {err['verification']}"""
        # 🔑 修复：移出多行字符串包裹，保证追加逻辑被正常物理执行
        unified_candidates.append(("ERR", weight, block_text))

    # 2. 压入关联指导准则 (GL)
    for gl in matched_guidelines:
        inherited_sev = gl.get("_inherited_severity", "MINOR")
        weight = SEVERITY_WEIGHT.get(inherited_sev, 1)
        block_text = f"""[ASSOCIATED GUIDELINE: {gl['id']} ({gl['category']})]
Guideline Fact: {gl['guideline']}
Target Scope: {", ".join(gl.get("target_files", ["*"]))}"""
        # 🔑 修复：移出多行字符串包裹，保证追加逻辑被正常物理执行
        unified_candidates.append(("GL", weight, block_text))

    # 3. 按照严重度序数权重由高到低进行最终排序
    unified_candidates.sort(key=lambda x: x[1], reverse=True)

    # 4. 抽取前 4 个对大模型最具威慑力、最底层的系统/编译级条目进行提取组装
    top_4_blocks = [item[2] for item in unified_candidates[:4]]
    final_rag_context = "\n\n".join(top_4_blocks)

    logger.info(
        f"--- [Few-shot RAG] Retrieved {len(matched_errors)} ERRs, {len(matched_guidelines)} GLs. Selected top 4 priorities. ---")

    return {
        "status": "success",
        "matched_errors_count": len(matched_errors),
        "associated_guidelines_count": len(matched_guidelines),
        "rag_context": final_rag_context
    }


def force_clean_git_repo(repo_path: str) -> Dict[str, str]:
    """
    Perform a deep clean of the specified Git repository with automated permission management.
    🔑 优化：引入原子权限回收工具，实现 1 行重用，精简了 20 多行样板代码；
    🔑 优化：统一在 cwd=abs_repo_path 绝对上下文中完成物理重置，避免目录频繁切换。
    """
    import os
    import subprocess
    print(f"--- Tool: force_clean_git_repo called for: {repo_path} ---")

    if not os.path.isdir(os.path.join(repo_path, ".git")):
        return {'status': 'error', 'message': f"'{repo_path}' is not a valid Git repository."}

    try:
        abs_repo_path = os.path.abspath(repo_path)

        # 1. 🔑 权限物理回收（替换掉原本复杂的 Docker/Host 跨层赋权及 fallback 自愈堆栈）
        reclaim_path_permissions(abs_repo_path)

        # 2. 统一在 cwd 上下文中完成物理重置与清理
        subprocess.run(["git", "reset", "--hard", "HEAD"], capture_output=True, text=True, check=True,
                       cwd=abs_repo_path)

        branch_res = subprocess.run(["git", "branch", "--list"], capture_output=True, text=True, cwd=abs_repo_path,
                                    check=True)
        main_branch = "main" if "main" in branch_res.stdout else "master"

        subprocess.run(["git", "switch", "-f", main_branch], capture_output=True, text=True, cwd=abs_repo_path,
                       check=True)
        subprocess.run(["git", "clean", "-fxd"], capture_output=True, text=True, cwd=abs_repo_path, check=True)

        return {'status': 'success', 'message': f"Successfully reclaimed and cleaned repo at '{repo_path}'."}
    except Exception as e:
        return {'status': 'error', 'message': f"Deep clean failed: {str(e)}"}



def extract_build_metadata_from_log(log_path: str) -> Dict:
    """
    Extract critical build metadata from the original error log using robust non-greedy regexes.
    Identifies GCR digest, compile flags, and third-party dependencies cleanly.
    """
    import os
    import re

    print(f"--- Tool: extract_build_metadata from {log_path} ---")
    try:
        if not os.path.exists(log_path):
            return {'status': 'error', 'message': 'Log file not found.'}

        with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        lines = content.splitlines()
        metadata = {
            'base_image_digest': '',
            'engine': 'libfuzzer',
            'sanitizer': 'address',
            'architecture': 'x86_64',
            'software_repo_url': '',
            'software_sha': '',
            'dependencies': []
        }

        # 🔑 1. 非贪婪 GCR base image digest 匹配提取
        digest_match = re.search(r'Digest:\s*?sha256:([a-f0-9]{64}?)', content, re.IGNORECASE)
        if digest_match:
            metadata['base_image_digest'] = digest_match.group(1)

        # 🔑 2. 非贪婪编译器流水线编译参数匹配
        for line in lines:
            if 'compile-' in line:
                m = re.search(r'compile-([a-z0-9]+)-([a-z0-9]+)-(x86_64|i386|arm64)', line)
                if m:
                    metadata['engine'], metadata['sanitizer'], metadata['architecture'] = m.groups()
                break

        # 🔑 3. 严格针对 Step #2 - "srcmap" 产生的依赖信息进行非贪婪提取
        git_pattern = re.compile(r'url:\s*?"([^"]+?)",\s*?rev:\s*?"([^"]+?)"')
        found_gits = []

        srcmap_zone = False
        for line in lines:
            if 'Step #2 - "srcmap"' in line:
                srcmap_zone = True
            elif 'Step #3 - "' in line:
                srcmap_zone = False

            if srcmap_zone:
                match = git_pattern.search(line)
                if match:
                    found_gits.append({'url': match.group(1), 'rev': match.group(2)})

        if found_gits:
            metadata['software_repo_url'] = found_gits[0]['url']
            metadata['software_sha'] = found_gits[0]['rev']
            metadata['dependencies'] = found_gits[1:]

        return {'status': 'success', 'metadata': metadata}
    except Exception as e:
        return {'status': 'error', 'message': f"Unexpected parser crash on log extraction: {e}"}


def patch_project_dockerfile(
        project_name: str,
        oss_fuzz_path: str,
        base_image_digest: str,
        dependencies: List[Dict] = None  # 新增参数：依赖列表
) -> Dict:
    import os
    import re

    print(f"--- Tool: patch_project_dockerfile (Enhanced) for {project_name} ---")
    dockerfile_path = os.path.join(oss_fuzz_path, "projects", project_name, "Dockerfile")

    try:
        with open(dockerfile_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # 1. 锁定 Base Image (现有逻辑)
        if base_image_digest:
            pattern = r'(FROM\s+gcr.io/oss-fuzz-base/base-builder[^\s:@]*)'
            content = re.sub(pattern + r'[^\s]*', rf'\1@sha256:{base_image_digest}', content, flags=re.IGNORECASE)

        # 2. 优化：移除所有 --depth 限制 (现有逻辑)
        content = re.sub(r'--depth[=\s]+?\d+', '', content, flags=re.IGNORECASE)

        # 3. 🔑 核心增强：注入第三方依赖的 git checkout 动作
        if dependencies:
            for dep in dependencies:
                url = dep.get('url', '')
                sha = dep.get('rev', '')
                if not url or not sha: continue

                # 💡 证据支撑优化：如果该依赖是主项目仓库，我们将通过宿主机物理挂载覆盖，无需在 Dockerfile 中对其执行 checkout
                # 提取仓库名称 (如 CRoaring)
                repo_name = url.split('/')[-1].replace('.git', '')
                if repo_name.lower() == project_name.lower():
                    print(f"  - Skip Dockerfile checkout injection for primary repo: {repo_name} (will be mounted)")
                    continue

                # 仅对次级依赖进行精准 checkout 注入
                pattern = rf"(git clone.*?{re.escape(repo_name)}[^\&\n;]*)"

                def inject_checkout(match):
                    original_line = match.group(1).strip()
                    tokens = original_line.split()
                    clone_dir = repo_name
                    try:
                        url_idx = -1
                        for idx, token in enumerate(tokens):
                            if repo_name in token:
                                url_idx = idx
                                break
                        if url_idx != -1 and url_idx + 1 < len(tokens):
                            next_token = tokens[url_idx + 1].strip()
                            if not next_token.startswith('-') and next_token not in ['&&', ';', '|']:
                                clone_dir = next_token
                    except Exception:
                        pass
                    return f"{original_line} && cd {clone_dir} && git checkout {sha} && cd .."

                content = re.sub(pattern, inject_checkout, content)

        with open(dockerfile_path, 'w', encoding='utf-8') as f:
            f.write(content)

        return {'status': 'success', 'message': "Dockerfile patched with pinned dependencies."}
    except Exception as e:
        return {'status': 'error', 'message': f'Failed to patch: {str(e)}'}


def update_yaml_report(file_path: str,
                       row_index: int,
                       result_str: str = None,
                       root_cause_commit: str = None,
                       root_cause_workspace: str = None) -> dict:
    """
    统一的 YAML 更新工具：支持可选的根因回写与最终状态更新，具备原子写入能力。
    """
    import os
    import yaml
    import tempfile
    from collections import OrderedDict
    from datetime import datetime

    try:
        if not os.path.exists(file_path):
            return {'status': 'error', 'message': f"YAML file not found: {file_path}"}

        with open(file_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)

        if row_index < 0 or row_index >= len(data):
            return {'status': 'error', 'message': f"Invalid row index: {row_index}"}

        entry = data[row_index]
        # 使用 OrderedDict 保持插入顺序
        new_entry = OrderedDict()
        for key, value in entry.items():
            new_entry[key] = value
            # 插入点：error_category 之后
            if key == 'error_category':
                if root_cause_commit and 'root_cause_commit' not in entry:
                    new_entry['root_cause_commit'] = root_cause_commit
                if root_cause_workspace and 'root_cause_workspace' not in entry:
                    new_entry['root_cause_workspace'] = root_cause_workspace

        # 如果提供了 result_str，更新状态
        if result_str:
            new_entry['fixed_state'] = 'no'
            new_entry['state'] = 'yes'
            new_entry['fix_result'] = result_str
            new_entry['fix_date'] = datetime.now().strftime("%Y-%m-%d")

        data[row_index] = dict(new_entry)

        # 原子写入逻辑
        dir_name = os.path.dirname(os.path.abspath(file_path))
        fd, tmp_path = tempfile.mkstemp(dir=dir_name, prefix=".yaml_tmp_", suffix=".yaml")
        try:
            with os.fdopen(fd, 'w', encoding='utf-8') as tmp_f:
                yaml.dump(data, tmp_f, default_flow_style=False, sort_keys=False, allow_unicode=True)
            os.replace(tmp_path, file_path)
        except Exception as e:
            if os.path.exists(tmp_path): os.remove(tmp_path)
            raise e

        return {'status': 'success', 'message': "YAML updated successfully."}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}


def get_git_commits_around_date(
        project_source_path: str,
        error_date: str,
        max_limit: int = 300,
        **kwargs  # Catch unexpected params
) -> Dict:
    """
    Retrieve ALL commits within a ±24h time window for comprehensive pre-screening.
    Optimized: Returns lightweight metadata (SHA/Date/Message) only.
    File changes & diffs are deferred to Phase 3 on-demand extraction to save time & tokens.
    """
    if 'count' in kwargs:
        raise ValueError("get_git_commits_around_date does not accept 'count' parameter. Use 'max_limit' instead.")

    if not ENABLE_HISTORY_ENHANCEMENT:
        print(f"--- [ABLATION] Temporal commit search is DISABLED. ---")
        return {'status': 'success', 'commits': [], 'total_count': 0}

    print(
        f"--- Tool: get_git_commits_around_date (Comprehensive Scan) | Path: {project_source_path} | Date: {error_date} ---")

    if not os.path.isdir(os.path.join(project_source_path, ".git")):
        return {'status': 'error', 'message': "Not a git repository."}

    try:
        # 容错解析日期
        target_dt = None
        if error_date and error_date.strip():
            for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d']:
                try:
                    target_dt = datetime.strptime(error_date.strip(), fmt)
                    break
                except ValueError:
                    continue

        if target_dt:
            start_date = (target_dt - timedelta(days=1)).strftime('%Y-%m-%d')
            end_date = (target_dt + timedelta(days=1)).strftime('%Y-%m-%d')
            print(f"--- Scanning commits between {start_date} and {end_date} (Limit: {max_limit}) ---")
            cmd = [
                "git", "log",
                f"--since={start_date} 00:00:00",
                f"--until={end_date} 23:59:59",
                f"--max-count={max_limit}",
                "--pretty=format:%H|%cd|%s",
                "--date=format:%Y-%m-%d %H:%M:%S"
            ]
        else:
            print(f"--- Date invalid. Falling back to recent {max_limit} commits. ---")
            cmd = ["git", "log", f"--max-count={max_limit}", "--pretty=format:%H|%cd|%s",
                   "--date=format:%Y-%m-%d %H:%M:%S"]

        result = subprocess.run(cmd, cwd=project_source_path, capture_output=True, text=True, check=False)

        commits = []
        for line in result.stdout.strip().split('\n'):
            if not line: continue
            parts = line.split('|', 2)
            if len(parts) < 3: continue
            sha, date, msg = parts
            # 🔑
            # 31/5000
            # Only return lightweight metadata. Do not trigger the git show query for file changes here.
            commits.append({
                "sha": sha,  # Full 40-char commit SHA
                "date": date,  # Formatted: YYYY-MM-DD HH:MM:SS
                "message": msg,  # First line of commit message (truncated if needed)
                "is_merge": msg.startswith("Merge"),  # Quick merge detection for Agent filtering
            })

        print(f"--- Found {len(commits)} commits in window. Ready for Agent pre-screening. ---")
        return {
            'status': 'success',
            'commits': commits,  # List[Dict{sha, date, message, is_merge}]
            'total_count': len(commits),
            'note': "File changes & diffs deferred to Phase 3 on-demand extraction via save_commit_diff_to_file/get_enhanced_history_context"
            # Help Agent understand workflow
        }
    except Exception as e:
        return {'status': 'error', 'message': f"Failed to get commits: {e}"}


def save_commit_diff_to_file(project_name: str, project_source_path: str, sha: str, error_time: str):
    """
    Extract recent changes and simplify based on content length to stay within token limits.
    """

    if not ENABLE_HISTORY_ENHANCEMENT:
        print(f"--- [ABLATION] Saving commit diff is DISABLED. ---")
        return {'status': 'error', 'message': 'History enhancement is disabled by ablation configuration.'}

    import os
    import subprocess
    print(f"--- Tool: save_commit_diff_to_file (With Token Guard) for {sha} ---")

    TOKEN_GUARD_CHARS = 12000
    OUTPUT_PATH = "generated_prompt_file/commit_changed.txt"
    os.makedirs("generated_prompt_file", exist_ok=True)

    try:
        raw_diff_res = subprocess.run(["git", "-C", project_source_path, "show", sha],
                                      capture_output=True, text=True, check=True)
        content = raw_diff_res.stdout

        if len(content) > TOKEN_GUARD_CHARS:
            print(f"  - Content length ({len(content)}) exceeds guard. Simplifying...")

            lines = content.split('\n')
            simplified = [l for l in lines if l.startswith(('+', '-', '@', 'commit', 'Author', 'Date'))]
            content = "\n".join(simplified)

            if len(content) > TOKEN_GUARD_CHARS:
                summary_res = subprocess.run(["git", "-C", project_source_path, "show", "--stat", sha],
                                             capture_output=True, text=True, check=True)
                content = "--- [DIFF TOO LARGE: Showing Summary Only] ---\n" + summary_res.stdout

        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            f.write(content)

        return {"status": "success", "message": f"Saved simplified diff to {OUTPUT_PATH}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def read_projects_from_yaml(file_path: str) -> dict:
    """
    Read project information, including state field checks and boolean compatibility.
    Exempts 'projects.yaml' from whitelist checks and catches syntax corruption safely.
    """
    import os
    import re
    import yaml
    from datetime import datetime
    from utils.path_utils import normalize_patch_path, validate_patch_path
    from utils.error_handler import format_path_error

    print(f"--- Tool: read_projects_from_yaml called for: {file_path} ---")

    # 🔑 1. 核心配置放行：projects.yaml 属于系统级核心引导文件，跳过子目录白名单限制
    if os.path.basename(file_path) == "projects.yaml":
        target_path = file_path
    else:
        normalized_path = normalize_patch_path(file_path)
        if not validate_patch_path(normalized_path):
            return {
                'status': 'error',
                'message': format_path_error(
                    original_path=file_path,
                    normalized_path=normalized_path,
                    base_dir=os.environ.get('PROJECT_ROOT', os.getcwd()),
                    validation_passed=False,
                    extra_info={'operation': 'read_projects_from_yaml'}
                )
            }
        target_path = normalized_path

    # 🔑 2. 安全检查：如果物理文件不存在，优雅返回错误字典
    if not os.path.exists(target_path):
        return {'status': 'error', 'message': f"YAML file not found at '{target_path}'."}

    projects_to_run = []
    try:
        # 🔑 3. 强固化异常捕获：预防 YAML 物理缩进损坏或格式错误引发 main() 闪退
        try:
            with open(target_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
        except yaml.YAMLError as ye:
            return {
                'status': 'error',
                'message': f"YAML parser failed (Syntax Error) in '{target_path}': {ye}"
            }
        except Exception as fe:
            return {
                'status': 'error',
                'message': f"Failed to open/read '{target_path}': {fe}"
            }

        if not isinstance(data, list):
            return {'status': 'error', 'message': f"YAML file '{target_path}' must contain a list of projects."}

        for index, entry in enumerate(data):
            fixed_state = str(entry.get('fixed_state', 'no')).lower()
            state = str(entry.get('state', 'no')).lower()

            if fixed_state == 'no' and state == 'no':
                project_name = entry.get('project')
                sha = entry.get('oss-fuzz_sha')
                error_time_str = str(entry.get('error_time', ""))
                fuzzing_build_error_log_url = entry.get('fuzzing_build_error_log', "")

                if project_name and sha:
                    log_dir = os.path.join("build_error_log", project_name)
                    original_log_path = ""

                    if fuzzing_build_error_log_url.startswith("http"):
                        from agent_tools import download_remote_log
                        download_result = download_remote_log(fuzzing_build_error_log_url, project_name, error_time_str)
                        if download_result['status'] == 'success':
                            original_log_path = download_result['local_path']

                    if not original_log_path and os.path.isdir(log_dir):
                        try:
                            y, m, d = map(int, error_time_str.replace('.', '-').split('-'))
                            base_date = datetime(y, m, d)
                            candidates = []
                            for filename in os.listdir(log_dir):
                                if "error.txt" in filename and re.match(r"\d{4}_\d{1,2}_\d{1,2} error\.txt", filename):
                                    match = re.search(r"(\d{4})_(\d{1,2})_(\d{1,2})", filename)
                                    if match:
                                        fy, fm, fd = map(int, match.groups())
                                        file_date = datetime(fy, fm, fd)
                                        if file_date >= base_date:
                                            candidates.append((file_date, filename))
                            if candidates:
                                candidates.sort(key=lambda x: x[0])
                                original_log_path = os.path.abspath(os.path.join(log_dir, candidates[0][1]))
                        except Exception:
                            pass

                    if original_log_path:
                        project_info = {
                            "project_name": project_name,
                            "sha": str(sha),
                            "row_index": index,
                            "error_time": error_time_str,
                            "original_log_path": original_log_path,
                            "software_repo_url": entry.get('software_repo_url', ""),
                            "software_sha": entry.get('software_sha', ""),
                            "engine": entry.get('engine', ""),
                            "sanitizer": entry.get('sanitizer', ""),
                            "architecture": entry.get('architecture', ""),
                            "base_image_digest": entry.get('base_image_digest', ""),
                            "root_cause_commit": entry.get('root_cause_commit', ""),
                            "root_cause_workspace": entry.get('root_cause_workspace', "")
                        }
                        projects_to_run.append(project_info)
                    else:
                        print(f"Warning: Project '{project_name}' skipped due to missing log file.")
                else:
                    print(f"Warning: Project at index {index} missing core fields. Skipping.")

        print(f"--- Found {len(projects_to_run)} projects to process. ---")
        return {'status': 'success', 'projects': projects_to_run}
    except Exception as e:
        return {'status': 'error', 'message': f"Unexpected failure in YAML parsing logic: {e}"}


def get_project_paths(project_name: str) -> Dict[str, str]:
    """
    Generates and returns the standard project_config_path and project_source_path based on the project name.
    """
    print(f"--- Tool: get_project_paths called for: {project_name} ---")
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__)))

    safe_project_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()

    config_path = os.path.join(base_path, "oss-fuzz", "projects", safe_project_name)
    config_repo_path = os.path.join(base_path, "oss-fuzz")
    source_path = os.path.join(base_path, "process", "project", safe_project_name)

    paths = {
        "project_name": project_name,
        "project_config_path": config_path,
        "project_config_repo_path": config_repo_path,
        "project_source_path": source_path,
        "max_depth": 1
    }
    print(f"--- Generated paths: {paths} ---")
    return paths


def get_workspace_root() -> Dict[str, str]:
    """
    Returns the absolute workspace root directory for dynamic path construction.
    """
    workspace_root = os.path.abspath(os.path.dirname(__file__))
    result = {"status": "success", "workspace_root": workspace_root}
    print(f"--- Tool: get_workspace_root called. Result: {result} ---")
    return result


def save_processed_project(project_name: str) -> Dict[str, str]:
    """
    Appends a processed project name to the project_processed.txt file.
    """
    print(f"--- Tool: save_processed_project called for: {project_name} ---")
    try:
        os.makedirs(PROCESSED_PROJECTS_DIR, exist_ok=True)
        with open(PROCESSED_PROJECTS_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{project_name}\n")
        message = f"Successfully saved '{project_name}' to processed list."
        print(f"--- {message} ---")
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"Failed to save processed project '{project_name}': {e}"
        print(f"--- ERROR: {message} ---")
        return {"status": "error", "message": message}


def update_excel_report(file_path: str, row_index: int, attempted: str, result: str) -> Dict[str, str]:
    """
    Updates the "Whether Fix Was Attempted", "Fix Result", and "Fix Date" columns for a specified row in an .xlsx file.
    """
    print(f"--- Tool: update_excel_report called for file '{file_path}', row {row_index} ---")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        attempted_col_idx = headers.index("是否尝试修复") + 1
        result_col_idx = headers.index("修复结果") + 1
        date_col_idx = headers.index("修复日期") + 1

        sheet.cell(row=row_index, column=attempted_col_idx, value=attempted)
        sheet.cell(row=row_index, column=result_col_idx, value=result)
        sheet.cell(row=row_index, column=date_col_idx, value=datetime.now().strftime('%Y-%m-%d'))

        workbook.save(file_path)
        message = f"Successfully updated row {row_index} in '{file_path}' with result: '{result}'."
        print(message)
        return {'status': 'success', 'message': message}
    except Exception as e:
        message = f"Failed to update Excel file: {e}"
        print(f"--- ERROR: {message} ---")
        return {'status': 'error', 'message': message}


def read_projects_from_excel(file_path: str) -> Dict:
    """
    Reads project information from the specified .xlsx file.
    Only reads rows where "Error Consistency" is "Yes" and "Whether Fix Was Attempted" is not "Yes".
    """
    print(f"--- Tool: read_projects_from_excel called for: {file_path} ---")
    if not os.path.exists(file_path):
        return {'status': 'error', 'message': f"Excel file not found at '{file_path}'."}

    projects_to_run = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        required_headers = ["项目名称", "复现oss-fuzz SHA", "报错是否一致", "是否尝试修复"]
        if not all(h in headers for h in required_headers):
            return {'status': 'error',
                    'message': f"Excel file is missing one of the required columns: {required_headers}"}

        name_idx = headers.index("项目名称")
        sha_idx = headers.index("复现oss-fuzz SHA")
        consistent_idx = headers.index("报错是否一致")
        attempted_idx = headers.index("是否尝试修复")

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[consistent_idx] == "是" and row[attempted_idx] != "是":
                project_info = {
                    "project_name": row[name_idx],
                    "sha": str(row[sha_idx]),
                    "row_index": row_index
                }
                projects_to_run.append(project_info)

        print(f"--- Found {len(projects_to_run)} new projects to process. ---")
        return {'status': 'success', 'projects': projects_to_run}
    except Exception as e:
        return {'status': 'error', 'message': f"Failed to read or parse Excel file: {e}"}


def run_command(command: str, timeout: int = 30, max_output_chars: int = 4000) -> dict:
    """
    Execute commands safely, compatible with LLM common Shell syntax,
    enforce zero-deletion policy, and return structured results.
    """
    print(f"--- Tool: run_command called with: '{command}' ---")

    # 🔒 强力零删除策略与高危拦截规则（使用词边界 \b 正则防止利用拼装、连写等手段绕过检测）
    # 明确禁止：任何文件/目录删除、权限篡改、系统级文件写入、远程网络下载、命令注入
    deletion_patterns = r'\b(?:rm|rmdir|unlink|del|shred|erase)\b'
    dangerous_patterns = r'\b(?:wget|curl|apt-get|apt|yum|sudo|su|chmod|chown|mkfs|dd|passwd|exec|eval)\b|>\s*/etc/|>\s*/var/|>\s*/sys/|\$\('

    if re.search(f'({deletion_patterns}|{dangerous_patterns})', command, re.IGNORECASE):
        return {
            "status": "error",
            "message": "🚫 Command blocked: Deletion or unsafe system-level operations are strictly forbidden. Use structured discovery tools instead (e.g., list_files_in_dir, read_file_content)."
        }

    try:
        # ✅ 使用 /bin/bash -c 兼容大模型常用 Shell 语法（如管道符 |、重定向 >、标准错误抑制等）
        res = subprocess.run(
            ['/bin/bash', '-c', command],
            capture_output=True, text=True, timeout=timeout, check=False
        )

        out = (res.stdout + res.stderr).strip()
        truncated = False
        if len(out) > max_output_chars:
            out = out[:max_output_chars] + f"\n[⚠️ OUTPUT TRUNCATED: {len(out) - max_output_chars} chars hidden]"
            truncated = True

        # 🎯 统合状态机语义：非零退出状态码显式标记为 error，降低 Agent 逻辑干扰
        return {
            "status": "success" if res.returncode == 0 else "error",
            "return_code": res.returncode,
            "output": out,
            "truncated": truncated,
            "hint": "Tip: Use `list_files_in_dir` for exploration, `read_file_content` for file inspection. Avoid complex shell chains." if res.returncode != 0 else ""
        }
    except subprocess.TimeoutExpired:
        return {
            "status": "error",
            "message": f"Command timed out after {timeout}s. Try `read_file_content` with mode='tail_N' or use `list_files_in_dir`."
        }
    except Exception as e:
        return {"status": "error", "message": f"Execution failed: {str(e)}"}


def truncate_prompt_file(file_path: str, max_lines: int = 2000) -> Dict[str, str]:
    """
    Reads a file, and if it exceeds max_lines, truncates it in the middle, keeping the head and tail.
    """
    print(f"--- Tool: truncate_prompt_file called for: {file_path} ---")
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        if len(lines) <= max_lines:
            message = "File is within line limits, no truncation needed."
            print(f"--- {message} ---")
            return {"status": "success", "message": message}

        head_count = max_lines // 4
        tail_count = max_lines - head_count

        truncated_content = "".join(lines[:head_count])
        truncated_content += "\n\n... (Content truncated due to context length limit) ...\n\n"
        truncated_content += "".join(lines[-tail_count:])

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(truncated_content)

        message = f"File '{file_path}' was truncated to approximately {max_lines} lines."
        print(f"--- {message} ---")
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"Failed to truncate file '{file_path}': {e}"
        print(f"--- ERROR: {message} ---")
        return {"status": "error", "message": message}


def archive_fixed_project(project_name: str, project_config_path: str, is_success: bool = True,
                          project_source_path: str = None) -> dict:
    import os, shutil, subprocess
    from datetime import datetime
    from agent_tools import TraceLedgerManager

    print(f"--- Tool: archive_fixed_project called for: {project_name} (Success: {is_success}) ---")
    try:
        # 1. 初始化路径与目录
        base_dir = "process/fixed" if is_success else "process/unfixed"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()
        destination_dir = os.path.join(os.getcwd(), base_dir, f"{safe_name}_{timestamp}")

        os.makedirs(destination_dir, exist_ok=True)  # 必须存在
        os.makedirs(os.path.join(destination_dir, "diffs"), exist_ok=True)

        # 2. 导出账本
        ledger_path = os.path.join(os.getcwd(), "project_repair_trace.json")
        sha_map = {}
        if os.path.exists(ledger_path):
            shutil.copy2(ledger_path, os.path.join(destination_dir, "project_repair_trace.json"))
            ledger = TraceLedgerManager.load_ledger()
            baseline = next((n for n in ledger.get("nodes", []) if n.get("node_id") == 0), {})
            sha_map = baseline.get("git_sha_state", {})

        # 2.1 纳入最终修复报告
        result_txt_path = os.path.join(os.getcwd(), "result.txt")
        if os.path.exists(result_txt_path):
            shutil.copy2(result_txt_path, os.path.join(destination_dir, "result.txt"))
            print(f"  - result.txt included in archive: {destination_dir}")
        else:
            print(f"  - Warning: result.txt not found at {result_txt_path}, skipping.")

        # 3. 统一提取变更 (Loop 结构)
        config_repo_path = project_config_path
        if config_repo_path and not os.path.exists(os.path.join(config_repo_path, ".git")):
            candidate_root = os.path.abspath(os.path.join(config_repo_path, "..", ".."))
            if os.path.exists(os.path.join(candidate_root, ".git")):
                config_repo_path = candidate_root

        targets = [
            (config_repo_path, "configs", "config_fix.patch", sha_map.get("oss-fuzz_sha")),
            (project_source_path, "source", "source_fix.patch", sha_map.get("project_sha"))
        ]
        
        for path, dest_sub, patch_name, baseline_sha in targets:
            if not path or not os.path.isdir(path): continue

            changed_files = []
            if baseline_sha and baseline_sha != "N/A":
                try:
                    res = subprocess.run(
                        ["git", "-C", path, "diff", "--name-only", "--diff-filter=ACMRT", baseline_sha, "HEAD"],
                        capture_output=True, text=True, check=True)
                    changed_files = [f.strip() for f in res.stdout.split('\n') if f.strip()]
                except Exception as e:
                    print(f"  - Warning: Diff failed for {path}: {e}")

            if changed_files:
                if dest_sub != "source":
                    for f_rel in changed_files:
                        src, dst = os.path.join(path, f_rel), os.path.join(destination_dir, dest_sub, f_rel)
                        os.makedirs(os.path.dirname(dst), exist_ok=True)
                        shutil.copy2(src, dst)
                with open(os.path.join(destination_dir, "diffs", patch_name), "w") as pf:
                    subprocess.run(["git", "-C", path, "diff", baseline_sha, "HEAD"], stdout=pf, check=True)
            else:
                if dest_sub != "source":
                    shutil.copytree(path, os.path.join(destination_dir, f"{dest_sub}_all"), dirs_exist_ok=True)
        
        # 4. 强制物理清理
        def _safe_physical_remove(dir_path: str):
            if not dir_path or not os.path.exists(dir_path): return
            print(f"  - Post-process cleanup: {dir_path}")
            try:
                reclaim_path_permissions(dir_path)
                shutil.rmtree(dir_path, ignore_errors=True)
            except Exception as e:
                print(f"    [WARNING] Cleanup failed: {e}")

        _safe_physical_remove(project_source_path)
        if project_config_path:
            oss_fuzz_root = os.path.abspath(os.path.join(project_config_path, "..", ".."))
            if os.path.exists(os.path.join(oss_fuzz_root, ".git")):
                _safe_physical_remove(oss_fuzz_root)

        return {"status": "success", "archive_dir": destination_dir}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def download_github_repo(project_name: str, target_dir: str, repo_url: Optional[str] = None) -> Dict[str, str]:
    """
    Download a repository with path enforcement and full cloning.
    🔑 优化：全面封杀 --depth=1，保障全量克隆以支持后续 ECRCL SHA 的 checkout；
    🔑 优化：强力强制重定向，将非 oss-fuzz 的第三方仓库牢牢锁定在 process/project/ 路径下。
    """
    import json
    import time
    import subprocess
    import os
    import shutil

    current_work_dir = os.getcwd()

    # 🔑 1. 下游 oss-fuzz 与 上游开源项目路径强制路由锁
    if project_name == "oss-fuzz":
        final_target_dir = os.path.abspath(target_dir)
    else:
        safe_name = "".join(c for c in project_name if c.isalnum() or c in ('_', '-')).rstrip()
        final_target_dir = os.path.abspath(os.path.join(current_work_dir, "process", "project", safe_name))

        if os.path.abspath(target_dir) != final_target_dir:
            print(f"--- Path Security Enforcement: Redirecting download from {target_dir} to {final_target_dir} ---")

    print(f"--- Tool: download_github_repo called for '{project_name}' ---")

    # 🔑 2. 若本地已存在合法的全量 Git 仓库，则跳过下载直接放行
    if os.path.isdir(final_target_dir) and os.path.exists(os.path.join(final_target_dir, ".git")):
        if project_name == "oss-fuzz":
            try:
                # 先清理锁定文件
                subprocess.run(["rm", "-f", ".git/index.lock"], cwd=final_target_dir)
                # 强制拉取
                subprocess.run(["git", "fetch", "origin"], cwd=final_target_dir, check=True)
                subprocess.run(["git", "reset", "--hard", "origin/master"], cwd=final_target_dir, check=True)
                return {'status': 'success', 'path': final_target_dir, 'message': 'oss-fuzz synced.'}
            except Exception as e:
                return {'status': 'success', 'path': final_target_dir, 'message': f'Sync failed: {e}'}

        else:
            print(f"--- Repo '{project_name}' exists and is a valid git repo. Skipping download. ---")
            return {'status': 'success', 'path': final_target_dir, 'message': 'Repository already exists.'}

    # 3. 准备物理清理
    if os.path.isdir(final_target_dir):
        safe_delete_path(final_target_dir)
    os.makedirs(os.path.dirname(final_target_dir), exist_ok=True)

    # 4. 远程 Repo URL 检索
    final_repo_url = repo_url if repo_url and repo_url.strip() else None
    if not final_repo_url:
        if project_name == "oss-fuzz":
            final_repo_url = "https://github.com/google/oss-fuzz.git"
        else:
            try:
                search_cmd = ["gh", "search", "repos", project_name, "--sort", "stars", "--limit", "1", "--json",
                              "fullName"]
                result = subprocess.run(search_cmd, capture_output=True, text=True, check=True, encoding='utf-8')
                parsed = json.loads(result.stdout.strip())
                if parsed:
                    final_repo_url = f"https://github.com/{parsed[0]['fullName']}.git"
                else:
                    return {'status': 'error', 'message': f"Repo not found for {project_name}"}
            except Exception as e:
                return {'status': 'error', 'message': f"Search failed: {e}"}

    # 5. Git 大文件传输参数优化
    subprocess.run(["git", "config", "--global", "http.postBuffer", "524288000"])
    subprocess.run(["git", "config", "--global", "http.lowSpeedLimit", "0"])
    subprocess.run(["git", "config", "--global", "http.lowSpeedTime", "999999"])

    max_retries = 3
    for attempt in range(max_retries):
        print(f"--- Download attempt {attempt + 1}/{max_retries} ---")
        try:
            # 🔑 6. 全量 clone 锁死（彻底抛弃 --depth 1 以免丢失提交树）
            clone_cmd = ["git", "clone", final_repo_url, final_target_dir]
            result = subprocess.run(clone_cmd, capture_output=True, text=True)
            if result.returncode == 0:
                return {'status': 'success', 'path': final_target_dir,
                        'message': 'Successfully cloned full history repository.'}
            else:
                print(f"--- Attempt {attempt + 1} failed: {result.stderr} ---")
        except Exception as e:
            print(f"--- Attempt {attempt + 1} exception: {e} ---")
        time.sleep(10 * (attempt + 1))

    return {'status': 'error', 'message': f"Failed to download {project_name} after {max_retries} attempts."}


def find_sha_for_timestamp(commits_file_path: str, error_date: str) -> Dict[str, str]:
    """
    Finds the most suitable commit SHA for a given date from a commits file.
    """
    print(f"--- Tool: find_sha_for_timestamp called for date: {error_date} ---")
    try:
        target_date = datetime.strptime(error_date, '%Y.%m.%d').date()
    except ValueError:
        return {'status': 'error', 'message': f"Invalid target date format: '{error_date}'. Expected 'YYYY.MM.DD'."}

    todays_commits: List[Tuple[datetime, str]] = []
    past_commits: List[Tuple[datetime, str]] = []

    try:
        with open(commits_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if line.startswith("Time: ") and i + 1 < len(lines) and lines[i + 1].strip().startswith("- SHA: "):
                try:
                    timestamp_str = line.replace("Time: ", "")
                    commit_datetime = datetime.strptime(timestamp_str, '%Y.%m.%d %H:%M')
                    sha = lines[i + 1].strip().replace("- SHA: ", "")
                    commit_date = commit_datetime.date()
                    if commit_date == target_date:
                        todays_commits.append((commit_datetime, sha))
                    elif commit_date < target_date:
                        past_commits.append((commit_datetime, sha))
                except (ValueError, IndexError):
                    pass
            i += 1
    except FileNotFoundError:
        return {'status': 'error', 'message': f"Commits file not found at: {commits_file_path}"}
    except Exception as e:
        return {'status': 'error', 'message': f"An unexpected error occurred: {e}"}

    if todays_commits:
        earliest_today = min(todays_commits)
        found_sha = earliest_today[1]
        return {'status': 'success', 'sha': found_sha}
    elif past_commits:
        latest_in_past = max(past_commits)
        found_sha = latest_in_past[1]
        return {'status': 'success', 'sha': found_sha}
    else:
        return {'status': 'error', 'message': f"No suitable SHA found on or before the date {error_date}."}


def checkout_oss_fuzz_commit(sha: str) -> Dict[str, str]:
    """
    Executes a git checkout command in the fixed oss-fuzz directory.
    🔑 优化：完全移除 os.chdir，利用 subprocess.run(cwd=...) 避免 CWD 漂移。
    """
    import os
    import subprocess

    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__)))
    oss_fuzz_path = os.path.join(base_path, "oss-fuzz")
    print(f"--- Tool: checkout_oss_fuzz_commit called for SHA: {sha} in '{oss_fuzz_path}' ---")

    if not os.path.isdir(os.path.join(oss_fuzz_path, ".git")):
        return {'status': 'error', 'message': f"The directory '{oss_fuzz_path}' is not a git repository."}

    try:
        # 在指定 oss_fuzz_path 路径下获取主干分支名
        branch_res = subprocess.run(["git", "branch"], capture_output=True, text=True, cwd=oss_fuzz_path, check=True)
        main_branch = "main" if "main" in branch_res.stdout else "master"

        # 强制切回主干并拉取
        subprocess.run(["git", "switch", main_branch], capture_output=True, text=True, cwd=oss_fuzz_path, check=True)

        # 执行目标 commit 的切换
        command = ["git", "checkout", sha]
        result = subprocess.run(command, capture_output=True, text=True, encoding='utf-8', cwd=oss_fuzz_path)

        if result.returncode == 0:
            return {'status': 'success', 'message': f"Successfully checked out SHA {sha} in oss-fuzz."}
        else:
            return {'status': 'error', 'message': f"Git checkout command failed in oss-fuzz: {result.stderr.strip()}"}
    except Exception as e:
        return {'status': 'error', 'message': f"An unexpected error occurred during oss-fuzz checkout: {e}"}


def save_file_tree(directory_path: str, output_file: Optional[str] = None) -> dict:
    """
    Gets the file tree structure of a specified directory path and saves it to a file.
    """
    print(f"--- Tool: save_file_tree called for path: {directory_path} ---")
    if not os.path.isdir(directory_path):
        error_message = f"Error: The provided path '{directory_path}' is not a valid directory."
        print(error_message)
        return {"status": "error", "message": error_message}
    if output_file is None:
        output_dir = "generated_prompt_file"
        final_output_path = os.path.join(output_dir, "file_tree.txt")
    else:
        final_output_path = output_file
    output_dir = os.path.dirname(final_output_path)
    try:
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        tree_lines = []

        def _build_tree_recursive(path, prefix=""):
            entries = sorted([e for e in os.listdir(path) if not e.startswith('.')])
            pointers = ["├── "] * (len(entries) - 1) + ["└── "]
            for pointer, entry in zip(pointers, entries):
                full_path = os.path.join(path, entry)
                if os.path.isdir(full_path):
                    tree_lines.append(f"{prefix}{pointer}📁 {entry}")
                    extension = "│   " if pointer == "├── " else "    "
                    _build_tree_recursive(full_path, prefix + extension)
                else:
                    tree_lines.append(f"{prefix}{pointer}📄 {entry}")

        tree_lines.insert(0, f"📁 {os.path.basename(os.path.abspath(directory_path))}")
        _build_tree_recursive(directory_path, prefix="")
        with open(final_output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(tree_lines))
        success_message = f"File tree has been successfully generated and saved to '{final_output_path}'."
        print(success_message)
        return {"status": "success", "message": success_message}
    except Exception as e:
        error_message = f"An error occurred while generating or saving the file tree: {str(e)}"
        print(error_message)
        return {"status": "error", "message": error_message}


@_safe_path_wrapper("save_file_tree_shallow")
def save_file_tree_shallow(directory_path: str, max_depth: int, output_file: Optional[str] = None, **kwargs) -> dict:
    """
    Generates a shallow file tree of the target directory.
    🔑 优化：强制接入 @_safe_path_wrapper 装饰器安全网；
    🔑 优化：限制一级根目录展示行数（最大 30 行），避免超大型开源项目撑爆 Agent 上下文。
    """
    import os

    if not os.path.isdir(directory_path):
        return {"status": "error", "message": f"Directory not found: {directory_path}"}

    if output_file is None:
        output_file = "generated_prompt_file/file_tree.txt"

    try:
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        tree_lines = []

        def _build_tree_recursive(path, prefix="", depth=0):
            if depth >= max_depth:
                return
            try:
                entries = sorted([e for e in os.listdir(path) if not e.startswith('.')])
            except OSError:
                entries = []

            # 🔑 1. 根目录与子目录级联配额控制：根目录最大展示 30 行，次级目录最大展示 15 行
            is_root = (depth == 0)
            limit = 30 if is_root else 15
            truncated = len(entries) > limit
            active_entries = entries[:limit]

            pointers = ["├── "] * (len(active_entries) - 1) + ["└── "]
            for pointer, entry in zip(pointers, active_entries):
                full_path = os.path.join(path, entry)
                if os.path.isdir(full_path) and not os.path.islink(full_path):
                    tree_lines.append(f"{prefix}{pointer}📁 {entry}")
                    extension = "│   " if pointer == "├── " else "    "
                    _build_tree_recursive(full_path, prefix + extension, depth + 1)
                else:
                    tree_lines.append(f"{prefix}{pointer}📄 {entry}")

            # 🔑 2. 提供可视化的截断提示，告知 Agent 当前结构被截断
            if truncated:
                tree_lines.append(f"{prefix}└── ... [truncated: {len(entries) - limit} entries hidden]")

        tree_lines.append(f"📁 {os.path.basename(os.path.abspath(directory_path))}")
        _build_tree_recursive(directory_path, prefix="", depth=0)

        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(tree_lines))

        return {"status": "success", "message": f"Shallow file tree saved successfully to {output_file}."}
    except Exception as e:
        return {"status": "error", "message": f"Failed to save file tree cleanly: {str(e)}"}


@_safe_path_wrapper("find_and_append_file_details")
def find_and_append_file_details(
        directory_path: str,
        search_keyword: str,
        output_file: Optional[str] = None,
        **kwargs
) -> dict:
    """
    Finds a file or directory matching a keyword and appends details.
    🔑 优化：强制接入 @_safe_path_wrapper 装饰器安全网；
    🔑 优化：加入物理真实路径(os.path.realpath)防循环死锁校验，剔除无限递归崩溃。
    """
    import os

    if not os.path.isdir(directory_path):
        return {"status": "error", "message": f"Directory not found: {directory_path}"}

    if output_file is None:
        output_file = "generated_prompt_file/file_tree.txt"

    try:
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        found_paths = []
        visited_real_paths = set()  # 🔑 用于防止无限递归的符号链接物理查重集合

        for root, dirs, files in os.walk(directory_path, followlinks=True):
            # 🔑 1. 计算并查重物理真实路径，如果是环形链接，直接修剪分支，抛弃子目录搜索
            real_root = os.path.realpath(root)
            if real_root in visited_real_paths:
                dirs.clear()  # 阻断 os.walk 向下级递归
                continue
            visited_real_paths.add(real_root)

            # 🔑 2. 安全截断目录列表，物理上剔除所有软链接目录
            dirs[:] = [d for d in dirs if not os.path.islink(os.path.join(root, d))]

            all_entries = dirs + files
            for entry in all_entries:
                full_path = os.path.join(root, entry)
                if search_keyword in full_path:
                    found_paths.append(full_path)

        found_paths = sorted(list(set(found_paths)))
        if not found_paths:
            message = f"No file or directory matching '{search_keyword}' was found."
            with open(output_file, "a", encoding="utf-8") as f:
                f.write(f"\n\n--- Detailed query result for '{search_keyword}' ---\n{message}\n")
            return {"status": "success", "message": message}

        details_to_append = [f"\n\n--- Detailed query result for '{search_keyword}' ---"]
        for path in found_paths:
            relative_path = os.path.relpath(path, directory_path)
            details_to_append.append(f"\n# Matched path: {relative_path}")
            if os.path.isdir(path):
                # 约束单个子目录展示数量，防止打印爆炸
                entries = sorted([e for e in os.listdir(path) if not e.startswith('.')])
                for entry in entries[:15]:
                    details_to_append.append(f"├── {'📁' if os.path.isdir(os.path.join(path, entry)) else '📄'} {entry}")
            else:
                details_to_append.append(f"📄 {os.path.basename(path)}")

        with open(output_file, "a", encoding="utf-8") as f:
            f.write("\n".join(details_to_append))

        return {"status": "success", "message": f"Appended details of '{search_keyword}' successfully."}
    except Exception as e:
        return {"status": "error", "message": f"Error in find_and_append_file_details: {str(e)}"}


@_safe_path_wrapper(operation_name="read_file_content")
def read_file_content(file_path: str, mode: str = "full", base_dir: str = None) -> dict:
    """
    【新旧结合型：高鲁棒性文件读取工具】
    1. 集成新机制：路径规范化、白名单校验、物理缺失路径纠错引导。
    2. 集成旧系统：License 自动剥离、多模式切片、500行硬熔断阈值防止 Token 溢出。
    """
    import os, re
    from utils.path_utils import DEFAULT_PROJECT_ROOT

    # 1. 路径解析基准对齐
    if base_dir is None:
        base_dir = DEFAULT_PROJECT_ROOT

    # 规范相对/绝对文件路径
    resolved_path = os.path.normpath(os.path.join(base_dir, file_path)) if not os.path.isabs(file_path) else file_path
    print(f"--- Tool: read_file_content (Mode: {mode}) called for: {file_path} (Resolved: {resolved_path}) ---")

    # 2. 物理文件缺失自愈：吐出纠错引导
    if not os.path.exists(resolved_path):
        from utils.error_handler import format_path_error
        error_guide = format_path_error(
            original_path=file_path,
            normalized_path=file_path,
            base_dir=base_dir,
            validation_passed=True,
            extra_info={
                'error': 'The whitelisted path is valid, but the target file does not physically exist on disk.'}
        )
        return {
            "status": "error",
            "message": f"File not found on disk:\n{error_guide}"
        }

    # 3. 执行读取与防御性处理
    try:
        with open(resolved_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        # A. 自动剥离 License/Header 头部 (节省 Token)
        license_pattern = re.compile(r"^(#|//|\s*\*|/\*).*$", re.MULTILINE)
        start_idx = 0
        for i, line in enumerate(lines[:50]):
            if line.strip() and not license_pattern.match(line):
                start_idx = i
                break
        if start_idx > 5:
            lines = lines[start_idx:]
            print(f"--- Stripped license header ({start_idx} lines) ---")

        total_lines = len(lines)
        MAX_SAFE_LINES = 500  # 硬熔断阈值

        # B. 根据模式进行按行切片，并执行 Safety Melt 策略
        if mode == "full":
            if total_lines > 1000:
                print(f"--- [SAFETY MELT] Full mode exceeds 1000 lines. Truncating to tail_{MAX_SAFE_LINES}. ---")
                lines = lines[-MAX_SAFE_LINES:]
                mode = "full (melted to tail_500)"
            else:
                pass  # 保留全文

        elif mode == "tail_50":  # 读取后 50%
            target_count = int(total_lines * 0.5)
            if target_count > MAX_SAFE_LINES:
                lines = lines[-100:]
                mode = "tail_50 (melted to 100)"
            else:
                lines = lines[-target_count:]

        elif mode == "tail_30":  # 读取后 30%
            target_count = int(total_lines * 0.3)
            if target_count > MAX_SAFE_LINES:
                lines = lines[-100:]
                mode = "tail_30 (melted to 100)"
            else:
                lines = lines[-target_count:]

        elif mode == "tail_200_lines":
            lines = lines[-200:]
        elif mode == "tail_100_lines":
            lines = lines[-100:]
        elif mode == "tail_40_lines":
            lines = lines[-40:]
        elif mode == "tail_50":  # 这里是固定的 50 行
            lines = lines[-50:]
        elif mode == "tail_30":  # 这里是固定的 30 行
            lines = lines[-30:]
        elif mode == "head_50":
            lines = lines[:50]
        else:
            # 未知模式默认硬截断为后 100 行
            if total_lines > 100:
                lines = lines[-100:]
                mode = f"unknown_{mode} (fallback to tail_100)"

        content = "".join(lines)
        return {
            "status": "success",
            "message": f"Read {len(lines)} lines from {file_path} (Effective Mode: {mode})",
            "content": content
        }

    except Exception as e:
        return {"status": "error", "message": f"Read operation failed: {str(e)}"}


@_safe_path_wrapper(operation_name="create_or_update_file")
def create_or_update_file(file_path: str, content: str, **kwargs) -> dict:
    """
    Creates a new file and writes content to it, or overwrites an existing file.
    Path normalization and whitelist validation are pre-checked by @_safe_path_wrapper.
    """
    from utils.path_utils import DEFAULT_PROJECT_ROOT

    # 路径解析基准对齐
    base_dir = kwargs.get('base_dir', os.environ.get('PROJECT_ROOT', DEFAULT_PROJECT_ROOT))
    resolved_path = os.path.normpath(os.path.join(base_dir, file_path)) if not os.path.isabs(file_path) else file_path

    print(f"--- Tool: create_or_update_file called for path: {file_path} (Resolved: {resolved_path}) ---")

    try:
        # 安全级联创建可能缺失的父目录结构
        directory = os.path.dirname(resolved_path)
        if directory:
            os.makedirs(directory, exist_ok=True)

        with open(resolved_path, "w", encoding="utf-8") as f:
            f.write(content)

        message = f"File '{file_path}' has been successfully created/updated."
        print(message)
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"An error occurred while creating or updating file '{file_path}': {str(e)}"
        print(message)
        return {"status": "error", "message": message}


def append_file_to_file(
        source_path: str,
        destination_path: str,
        base_dir: Optional[str] = None,
        strict_mode: bool = True
) -> dict:
    """
    Reads the entire content of a source file and appends it to the end of a destination file.
    Optimized: Path normalization + whitelist validation for both paths.
    """
    from utils.path_utils import normalize_patch_path, validate_patch_path
    from utils.error_handler import format_path_error

    print(f"--- Tool: append_file_to_file called. Source: '{source_path}', Destination: '{destination_path}' ---")

    if base_dir is None:
        base_dir = os.environ.get('PROJECT_ROOT', os.getcwd())

    # 🔐 双路径校验
    normalized_source = normalize_patch_path(source_path, base_dir)
    normalized_dest = normalize_patch_path(destination_path, base_dir)

    if strict_mode:
        if not validate_patch_path(normalized_source, strict=True):
            return {
                "status": "error",
                "message": format_path_error(
                    original_path=source_path,
                    normalized_path=normalized_source,
                    base_dir=base_dir,
                    validation_passed=False,
                    extra_info={'operation': 'append_file_to_file', 'path_type': 'source'}
                )
            }
        if not validate_patch_path(normalized_dest, strict=True):
            return {
                "status": "error",
                "message": format_path_error(
                    original_path=destination_path,
                    normalized_path=normalized_dest,
                    base_dir=base_dir,
                    validation_passed=False,
                    extra_info={'operation': 'append_file_to_file', 'path_type': 'destination'}
                )
            }

    if not os.path.isfile(normalized_source):
        return {"status": "error",
                "message": f"Error: Source file '{source_path}' does not exist or is not a valid file."}
    if os.path.isdir(normalized_dest):
        return {"status": "error",
                "message": f"Error: Destination path '{destination_path}' is a directory and cannot be an append target."}
    if os.path.abspath(normalized_source) == os.path.abspath(normalized_dest):
        return {"status": "error", "message": "Error: Source and destination files cannot be the same."}

    try:
        with open(normalized_source, "r", encoding="utf-8") as f_source:
            content_to_append = f_source.read()
        dest_directory = os.path.dirname(normalized_dest)
        if dest_directory:
            os.makedirs(dest_directory, exist_ok=True)
        with open(normalized_dest, "a", encoding="utf-8") as f_dest:
            f_dest.write(content_to_append)
        return {"status": "success",
                "message": f"Successfully appended the content of '{source_path}' to '{destination_path}'."}
    except Exception as e:
        return {"status": "error", "message": f"An unknown error occurred while appending the file: {str(e)}"}


def append_string_to_file(
        file_path: str,
        content: str,
        base_dir: Optional[str] = None,
        strict_mode: bool = True
) -> dict:
    """
    Appends a string of content to the end of a specified file.
    Optimized: Path normalization + whitelist validation.
    """
    from utils.path_utils import normalize_patch_path, validate_patch_path
    from utils.error_handler import format_path_error

    print(f"--- Tool: append_string_to_file called for path: {file_path} ---")

    if base_dir is None:
        base_dir = os.environ.get('PROJECT_ROOT', os.getcwd())

    normalized_path = normalize_patch_path(file_path, base_dir)
    if strict_mode and not validate_patch_path(normalized_path, strict=True):
        return {
            "status": "error",
            "message": format_path_error(
                original_path=file_path,
                normalized_path=normalized_path,
                base_dir=base_dir,
                validation_passed=False,
                extra_info={'operation': 'append_string_to_file'}
            )
        }

    try:
        directory = os.path.dirname(normalized_path)
        if directory:
            os.makedirs(directory, exist_ok=True)
        with open(normalized_path, "a", encoding="utf-8") as f:
            f.write(content)
        return {"status": "success", "message": f"Content successfully appended to file '{file_path}'."}
    except Exception as e:
        return {"status": "error",
                "message": f"An error occurred while appending content to file '{file_path}': {str(e)}"}


def delete_file(file_path: str, base_dir: str = None, **kwargs) -> dict:
    """
    Deletes a specified file.
    """
    if base_dir is None:
        base_dir = os.environ.get('PROJECT_ROOT', 'fix_build_agent')

    normalized_path = normalize_patch_path(file_path, base_dir)

    # 白名单验证（删除操作必须严格）
    strict_mode = kwargs.get('strict_mode', True)
    if strict_mode and not validate_patch_path(normalized_path, strict=True):
        return {
            "status": "error",
            "message": format_path_error(
                original_path=file_path,
                normalized_path=normalized_path,
                base_dir=base_dir,
                validation_passed=False,
                extra_info={'operation': 'delete_file'}
            )
        }
    print(f"--- Tool: delete_file called for path: {normalized_path} ---")
    if not os.path.exists(normalized_path):
        message = f"Error: File '{file_path}' does not exist and cannot be deleted."
        print(message)
        return {"status": "error", "message": message}
    try:
        safe_delete_path(normalized_path)
        message = f"File '{normalized_path}' has been successfully deleted."
        print(message)
        return {"status": "success", "message": message}
    except Exception as e:
        message = f"An error occurred while deleting file '{file_path}': {str(e)}"
        print(message)
        return {"status": "error", "message": message}


def prompt_generate_tool(
        tool_context: ToolContext,
        project_main_folder_path: str,
        max_depth: int,
        config_folder_path: str,
        attempt_id: int
) -> dict:
    """
    物理组装工具：从账本、归因工件和RAG库中提取信息并生成最终的 prompt.txt。
    """
    import os, re

    print(f"--- Workflow Tool: prompt_generate_tool started (Attempt: {attempt_id}) ---")

    session = tool_context.session
    current_node_id = session.state.get("current_node_id", 0)
    validation_report = session.state.get("last_validation_report", {})
    basic_info = extract_basic_information(session.state.get("basic_information"))
    if basic_info.get("project_source_path"):
        project_main_folder_path = basic_info["project_source_path"]
    if basic_info.get("project_config_path"):
        config_folder_path = basic_info["project_config_path"]
    print(
        "[DEBUG prompt_generate context] "
        f"project_name={session.state.get('project_name')} | "
        f"project_source_path={session.state.get('project_source_path')} | "
        f"project_config_path={session.state.get('project_config_path')} | "
        f"project_config_repo_path={session.state.get('project_config_repo_path')} | "
        f"basic_information={session.state.get('basic_information')} | "
        f"project_main_folder_path={project_main_folder_path} | "
        f"config_folder_path={config_folder_path}"
    )

    PROMPT_DIR = "generated_prompt_file"
    PROMPT_FILE_PATH = os.path.abspath(os.path.join(PROMPT_DIR, "prompt.txt"))
    FUZZ_LOG_PATH = "fuzz_build_log_file/fuzz_build_log.txt"
    os.makedirs(PROMPT_DIR, exist_ok=True)

    # =================================================================
    # 1. 自动组装历史策略轨迹 (自适应节点检查)
    # =================================================================
    enhanced_history = ""

    if current_node_id == 0:
        enhanced_history = "LOG: This is the initial baseline attempt. No previous repair history exists."
    else:
        ledger = TraceLedgerManager.load_ledger()
        parent_chain_ids = _collect_parent_chain_node_ids(ledger, current_node_id, limit=3, include_self=False)
        history_labels = ["ROUND N", "ROUND N-1", "ROUND N-2"]
        for idx, nid in enumerate(parent_chain_ids):
            res = query_trace_ledger(
                tool_context=tool_context,
                field_keys=[
                    "action_and_intent.repair_strategy",
                    "action_and_intent.loop_summary",
                    "semantic_memory.reflection_analysis"
                ],
                node_id=nid
            )
            if res["status"] == "success":
                data = res["data"]
                label = "INITIAL BASELINE" if nid == 0 else history_labels[idx] if idx < len(history_labels) else f"ROUND {nid}"
                enhanced_history += f"\n--- [{label} REFLECTION] ---\n"
                enhanced_history += f"Strategy: {data.get('action_and_intent.repair_strategy', 'N/A')}\n"
                enhanced_history += f"Summary: {data.get('action_and_intent.loop_summary', 'N/A')}\n"
                enhanced_history += f"Reflection: {data.get('semantic_memory.reflection_analysis', 'N/A')}\n"

    if not enhanced_history.strip():
        enhanced_history = "No relevant historical trajectory found in trace ledger."

    # =================================================================
    # 2. 提取当前 ECRCL 归因工件 (故障根因)
    # =================================================================
    causal_chain = "N/A"
    final_attribution = "N/A"
    commit_changed_path = os.path.abspath("generated_prompt_file/commit_changed.txt")

    if os.path.exists(commit_changed_path):
        try:
            with open(commit_changed_path, 'r', encoding='utf-8') as f:
                txt = f.read()
                # 🔑 Optimized Parser: Detect fallback mode
                if "[STATUS]: FAILED" in txt:
                    causal_chain = "Localization failed. System has automatically switched to Log-Based Diagnostic Mode."
                else:
                    cc_match = re.search(r"\[CAUSAL_CHAIN\]\s*([\s\S]*?)(?=\n\n\[|$)", txt)
                    if cc_match: causal_chain = cc_match.group(1).strip()

                fa_match = re.search(r"\[FINAL_ATTRIBUTION\]\s*([\s\S]*)$", txt)
                if fa_match: final_attribution = fa_match.group(1).strip()
        except Exception as e:
            print(f"Warning: Failed to parse commit_changed.txt: {e}")

    # =================================================================
    # 3. 物理触发 Few-shot RAG 检索
    # =================================================================
    rag_res = few_shot_rag_retrieve("expert_knowledge.json", FUZZ_LOG_PATH)
    expert_context = rag_res.get("rag_context", "No expert knowledge matched.")

    # =================================================================
    # 4. 组装最终 Prompt 文件
    # =================================================================
    project_name = os.path.basename(os.path.abspath(project_main_folder_path))

    with open(PROMPT_FILE_PATH, "w", encoding="utf-8") as f:
        f.write(f"Testing Expert. Project: {project_name}. Attempt: {attempt_id}\n")

        f.write("\n--- 【LAST BUILD VALIDATION (1+2+6 CRITERIA)】 ---\n")
        for k in ["step_1_official_list", "step_2_infra_compliance", "step_6_runtime_stability"]:
            f.write(f"{k.upper()}: {validation_report.get(k, 'N/A')}\n")

        f.write(f"\n【STRATEGIC KNOWLEDGE (RAG)】\n{expert_context}\n")
        f.write(f"\n【CAUSAL_CHAIN】\n{causal_chain}\n")
        f.write(f"\n【FINAL_ATTRIBUTION】\n{final_attribution}\n")
        f.write(f"\n【REPAIR_HISTORY_TRAJECTORY】\n{enhanced_history}\n")

        # 注入 Docker/Build 配置文件
        for fname in sorted(os.listdir(config_folder_path)):
            file_abs_path = os.path.join(config_folder_path, fname)
            if os.path.isfile(file_abs_path) and (
                    fname.endswith('.sh') or 'Dockerfile' in fname or 'project.yaml' in fname):
                # 显式使用 read_file_content 确保 Safety Melt 生效
                res_content = read_file_content(file_abs_path, mode="full")
                if res_content.get("status") == "success":
                    f.write(f"\n### {fname} ###\n{res_content.get('content', '')}\n")

        # 记录浅层文件树
        save_file_tree_shallow(project_main_folder_path, 1, os.path.join(PROMPT_DIR, "file_tree.txt"))

        # 注入日志尾部上下文 (严格限制长度)
        if os.path.exists(FUZZ_LOG_PATH):
            with open(FUZZ_LOG_PATH, 'r', encoding='utf-8', errors='ignore') as lf:
                # 只取最后 12000 字符，约 2000-3000 Token
                f.write(f"\n\n--- BUILD LOG TAIL ---\n{lf.read()[-12000:]}")

    # 5. 最终截断保护 (由外部配置决定阈值)
    limit_lines = globals().get("MAX_LINES_LIMIT", 2500)
    truncate_prompt_file(PROMPT_FILE_PATH, max_lines=limit_lines)

    return {"status": "success", "content": "Prompt successfully assembled."}


def _auto_discover_project_symbols(binary_path: str, project_name: str) -> Optional[List[str]]:
    """Heuristically identify project-specific symbols using nm."""
    import subprocess
    try:
        result = subprocess.run(['nm', '-D', binary_path], capture_output=True, text=True, errors='ignore')
        if result.returncode != 0:
            result = subprocess.run(['nm', binary_path], capture_output=True, text=True, errors='ignore')

        lines = result.stdout.splitlines()
        keywords = [project_name.lower(), "deflate", "inflate", "adler32", "crc32"] if project_name == "zlib" else [
            project_name.lower()]
        boilerplate = ('__asan', '__lsan', '__ubsan', '__sanitizer', 'fuzzer::', 'LLVM', 'afl_', '_Z', 'std::')

        candidates = []
        for line in lines:
            parts = line.split()
            if not parts: continue
            symbol = parts[-1]
            if any(kw in symbol.lower() for kw in keywords) and not symbol.startswith(boilerplate):
                candidates.append(symbol)
        return candidates[:5] if candidates else None
    except Exception as e:
        logger.debug(f"Cleanup step failed (non-fatal): {e}")


def _cleanup_environment(oss_fuzz_path: str, project_name: str):
    """
    全方位立体强杀：从项目镜像、Runner镜像、物理挂载卷三个维度彻底解除锁定。
    """
    import subprocess
    import os

    # 获取需要排查的宿主机物理挂载路径
    host_out_dir = os.path.join(oss_fuzz_path, "build", "out", project_name)

    try:
        # 1. 强杀并删除所有物理挂载了该输出目录的任何活跃容器（最稳健，直击痛点）
        if os.path.exists(host_out_dir):
            volume_filter = f"volume={host_out_dir}"
            subprocess.run(
                f"docker rm -f $(docker ps -a -q --filter {volume_filter}) 2>/dev/null",
                shell=True, capture_output=True, timeout=10
            )

        # 2. 强杀属于该项目的构建容器
        project_filter = f"ancestor=gcr.io/oss-fuzz/{project_name}"
        subprocess.run(
            f"docker rm -f $(docker ps -a -q --filter {project_filter}) 2>/dev/null",
            shell=True, capture_output=True, timeout=10
        )

        # 3. 强杀所有潜在残留的 base-runner 容器（带各种 Tag 的通配过滤）
        runner_filter = "ancestor=gcr.io/oss-fuzz-base/base-runner"
        subprocess.run(
            f"docker rm -f $(docker ps -a -q --filter {runner_filter}) 2>/dev/null",
            shell=True, capture_output=True, timeout=10
        )

        # 4. 后台残留进程安全扫描（如果 docker 层面清理完仍有锁，强杀对应宿主机残留影子进程）
        # 利用 pkill 杀掉可能脱离容器的宿主机 afl-fuzz 孤儿进程
        subprocess.run("pkill -9 -f afl-fuzz 2>/dev/null", shell=True, capture_output=True)

    except Exception as e:
        print(f"[*] Comprehensive cleanup encountered a warning: {e}")


def _auto_discover_project_symbols_from_content(nm_stdout: str, project_name: str) -> bool:
    """Helper to evaluate static linkage of project logic from symbol table."""
    keywords = [project_name.lower(), "deflate", "inflate", "adler32", "crc32"] if project_name == "zlib" else [
        project_name.lower()]
    boilerplate = ('__asan', '__lsan', '__ubsan', '__sanitizer', 'fuzzer::', 'LLVM', 'afl_', '_Z', 'std::')

    for line in nm_stdout.splitlines():
        parts = line.split()
        if not parts: continue
        symbol = parts[-1]
        if any(kw in symbol.lower() for kw in keywords) and not symbol.startswith(boilerplate):
            return True
    return False


@_safe_path_wrapper
def modify_file_by_lines(
        file_path: str,
        operation: str,
        line_number: int,
        end_line: Optional[int] = None,
        content: str = "",
        **kwargs
) -> dict:
    base_dir = kwargs.get('base_dir', os.getcwd())
    normalized_path = os.path.normpath(os.path.join(base_dir, file_path)) if not os.path.isabs(file_path) else file_path

    if not os.path.isfile(normalized_path):
        return {"status": "error", "message": "File not found."}
    valid_ops = {"insert_after", "insert_before", "delete", "replace"}
    if operation not in valid_ops:
        return {"status": "error", "message": f"Invalid operation. Must be one of {valid_ops}."}

    if end_line is None:
        end_line = line_number
    if line_number < 1 or end_line < line_number:
        return {"status": "error", "message": "Invalid line range."}

    with open(normalized_path, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()
    total_lines = len(lines)

    if line_number > total_lines:
        return {"status": "error", "message": f"Line exceeds file length."}
    if end_line > total_lines:
        end_line = total_lines

    start_idx = line_number - 1
    end_idx = end_line

    new_lines = []
    if operation == "delete":
        new_lines = lines[:start_idx] + lines[end_idx:]
    elif operation == "replace":
        if not content.endswith('\n'):
            content += '\n'
        new_lines = lines[:start_idx] + content.splitlines(keepends=True) + lines[end_idx:]
    elif operation.startswith("insert"):
        if not content.endswith('\n'):
            content += '\n'
        insert_pos = start_idx if operation == "insert_before" else end_idx
        new_lines = lines[:insert_pos] + content.splitlines(keepends=True) + lines[insert_pos:]

    try:
        with open(normalized_path, 'w', encoding='utf-8') as f:
            f.writelines(new_lines)
        return {"status": "success", "message": f"Applied '{operation}' at lines {line_number}-{end_line}."}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def list_files_in_dir(
        dir_path: str,
        max_depth: int = 2,
        pattern: str = "*",
        max_results: int = 200,
        base_dir: Optional[str] = None,
        strict_mode: bool = True
) -> dict:
    """
    Return a structured, LLM-friendly file tree. Replaces `run_command + find`.
    Optimized: Path normalization + symlink protection + whitelist validation.
    """
    if base_dir is None:
        base_dir = os.environ.get('PROJECT_ROOT', os.getcwd())

    normalized_dir = normalize_patch_path(dir_path, base_dir)
    if strict_mode and not validate_patch_path(normalized_dir, strict=True):
        return {
            "status": "error",
            "message": format_path_error(
                original_path=dir_path,
                normalized_path=normalized_dir,
                base_dir=base_dir,
                validation_passed=False,
                extra_info={'operation': 'list_files_in_dir'}
            )
        }

    if not os.path.isdir(normalized_dir):
        return {"status": "error", "message": "Directory not found."}

    results = []
    visited_real_paths = set()  # 🔐 防止符号链接循环

    def _traverse(current: str, depth: int):
        if depth > max_depth or len(results) >= max_results:
            return
        # 🔐 符号链接保护
        real_path = os.path.realpath(current)
        if real_path in visited_real_paths:
            return
        visited_real_paths.add(real_path)

        try:
            entries = sorted(os.listdir(current))
        except PermissionError:
            return

        for entry in entries:
            if len(results) >= max_results:
                break
            full_path = os.path.join(current, entry)
            rel_path = os.path.relpath(full_path, normalized_dir)

            if fnmatch.fnmatch(entry, pattern) or fnmatch.fnmatch(rel_path, f"*{pattern}*"):
                is_dir = os.path.isdir(full_path) and not os.path.islink(full_path)  # 🔐 排除符号链接目录
                results.append({"path": rel_path, "type": "dir" if is_dir else "file"})

            if os.path.isdir(full_path) and not os.path.islink(full_path):  # 🔐 不递归符号链接
                _traverse(full_path, depth + 1)

    _traverse(normalized_dir, 0)

    return {
        "status": "success",
        "count": len(results),
        "files": results[:max_results],
        "truncated": len(results) > max_results
    }


def check_file_exists(file_path: str) -> dict:
    """
    Safely checks if a file exists within the workspace.
    Replaces unsafe 'ls ... 2>/dev/null' shell commands with structured JSON response.
    """
    import os
    # 1. 路径安全规范化（防穿越）
    workspace_root = os.getcwd()
    target = os.path.normpath(
        os.path.join(workspace_root, file_path) if not os.path.isabs(file_path) else file_path
    )
    if not os.path.realpath(target).startswith(os.path.realpath(workspace_root)):
        return {"status": "error", "message": "Path validation failed: access denied."}

    # 2. 返回结构化状态
    return {
        "status": "success",
        "exists": os.path.isfile(target),
        "path": target
    }


class EvidenceGraph:
    def __init__(self):
        self.nodes = {}
        self.edges = {}

    def add_node(self, node_id: str, node_type: str):
        self.nodes[node_id] = node_type
        if node_id not in self.edges:
            self.edges[node_id] = {}

    def add_edge(self, u: str, v: str, weight: float):
        if u in self.nodes and v in self.nodes:
            self.edges[u][v] = weight

    def run_belief_propagation(self, active_commits: List[str], current_env: dict,
                               commit_messages: Optional[Dict[str, str]] = None) -> Dict[str, float]:
        d = 0.85
        Pt = {nid: 0.0 for nid in self.nodes}
        if "Nregion" in Pt: Pt["Nregion"] = 1.0
        in_neighbors = {nid: [] for nid in self.nodes}
        for u, out_edges in self.edges.items():
            sum_w = sum(out_edges.values())
            if sum_w > 0:
                for v, w in out_edges.items():
                    in_neighbors[v].append((u, w / sum_w))
        for t in range(3):
            next_Pt = {nid: 0.0 for nid in self.nodes}
            for u in self.nodes:
                I_u = 1.0 if u == "Nregion" else 0.0
                sum_in = sum(prob * Pt[v] for v, prob in in_neighbors[u])
                next_Pt[u] = (1.0 - d) * I_u + d * sum_in
            Pt = next_Pt
            env_san = current_env.get("SANITIZER", "address").lower()
            for commit in active_commits:
                commit_node = f"Ncommit_{commit}"
                if commit_node in Pt:
                    commit_msg = commit_messages.get(commit, "").lower() if commit_messages else ""
                    if (("asan" in env_san and "msan" in commit_msg) or ("msan" in env_san and "asan" in commit_msg)):
                        Pt[commit_node] = 0.0
        return {commit: Pt.get(f"Ncommit_{commit}", 0.0) for commit in active_commits}


def clamp_diff_content(diff_text: str) -> str:
    """
    Enforces token budget limits on Unified Diff content.
    1. Single file diffs over 500 chars are pruned to keep only hunk headers and modified lines.
    2. Total diff over 3000 chars is pruned to keep +, -, @ and headers.
    3. Exceeding 3000 chars after pruning keeps the current core-line-only fallback behavior.
    """
    if not diff_text:
        return ""

    # 按照文件划分 Diff 块
    file_blocks = []
    current_block = []
    for line in diff_text.splitlines():
        if line.startswith("diff --git "):
            if current_block:
                file_blocks.append("\n".join(current_block))
            current_block = [line]
        else:
            current_block.append(line)
    if current_block:
        file_blocks.append("\n".join(current_block))

    clamped_blocks = []
    for block in file_blocks:
        if len(block) > 500:
            # 强行截断：抛弃未修改的上下文行，仅保留 Hunk Header + 修改行
            lines = block.splitlines()
            pruned_lines = [
                l for l in lines
                if l.startswith(('+', '-', '@', 'diff --git ', '--- ', '+++ ', 'index '))
            ]
            clamped_block = "\n".join(pruned_lines)
            if len(clamped_block) > 500:
                clamped_block = clamped_block[:500] + "\n... [Single File Diff Truncated] ..."
            clamped_blocks.append(clamped_block)
        else:
            clamped_blocks.append(block)

    final_diff = "\n".join(clamped_blocks)

    if len(final_diff) > 3000:
        # 总 Diff 过滤，仅保留核心标记行
        lines = final_diff.splitlines()
        shrunk_lines = [
            l for l in lines
            if l.startswith(('+', '-', '@', 'diff --git ', 'commit ', 'Author:', 'Date:', 'Subject:'))
        ]
        final_diff = "\n".join(shrunk_lines)

    return final_diff


def timezone_normalize(error_date: str) -> int:
    """
    将多样化的时间格式 (CST UTC+8) 转换为 UTC 标准 naive Epoch 时间戳。
    确保在 Git 历史分析中时序对齐无偏差。
    """
    try:
        tz_cst = timezone(timedelta(hours=8))
        clean_date = error_date.strip().replace('.', '-').replace('/', '-')

        if ' ' in clean_date:
            t_error_naive = datetime.strptime(clean_date, "%Y-%m-%d %H:%M:%S")
        else:
            t_error_naive = datetime.strptime(clean_date, "%Y-%m-%d")

        t_error_cst = t_error_naive.replace(tzinfo=tz_cst)
        t_error_utc = t_error_cst.astimezone(timezone.utc)
        return int(t_error_utc.timestamp())
    except Exception as e:
        logger.warning(f"Failed to normalize error_date '{error_date}': {e}. Falling back to now.")
        return int(datetime.now(timezone.utc).timestamp())


def git_revert_counterfactual(repo_path: str, target_commit: str, project_name: str,
                              engine: str, sanitizer: str, architecture: str) -> bool:
    """
    下游commit反事实验证逻辑。
    作用：保存现场 -> 尝试revert -> 编译校验 -> 安全还原。
    修复安全缺陷：引入显式回滚异常拦截和基于 HEAD hash 的无损现场重置，杜绝抹除正常历史提交的隐患。
    """
    import subprocess
    # 1. 显式读取并保存当前的 HEAD 哈希，避免使用不确定的 HEAD~1
    orig_head = subprocess.run(
        ["git", "-C", repo_path, "rev-parse", "HEAD"],
        capture_output=True, text=True
    ).stdout.strip()

    # 使用 git stash 暂存所有工作区修改与未跟踪文件
    subprocess.run(["git", "-C", repo_path, "stash", "--include-untracked"], capture_output=True)
    try:
        # 2. 执行 revert 并截获冲突状态
        revert_res = subprocess.run(
            ["git", "-C", repo_path, "revert", "--no-edit", target_commit],
            capture_output=True
        )
        if revert_res.returncode != 0:
            # revert 过程若产生合并冲突，执行 abort 放弃冲突状态并安全退出
            subprocess.run(["git", "-C", repo_path, "revert", "--abort"], capture_output=True)
            return False

        # 3. 执行构建校验 (下游验证，mount_path = None)
        res = run_fuzz_build_and_validate(
            project_name=project_name,
            oss_fuzz_path=repo_path,
            sanitizer=sanitizer,
            engine=engine,
            architecture=architecture,
            mount_path=None
        )
        return res["status"] == "success"
    except Exception as e:
        logger.warning(f"Revert counterfactual hit exception on {target_commit}: {e}")
        return False
    finally:
        # 4. 无论中间是否发生异常，强制无损重置回我们保存好的 orig_head 物理锚点
        subprocess.run(["git", "-C", repo_path, "reset", "--hard", orig_head], capture_output=True)
        subprocess.run(["git", "-C", repo_path, "clean", "-fdx"], capture_output=True)
        subprocess.run(["git", "-C", repo_path, "stash", "pop"], capture_output=True)


def run_ecrcl_localization(
        log_path: str,
        project_name: str,
        project_source_path: str,
        oss_fuzz_path: str,
        error_date: str,
        suggested_find_commit_path: str = "",
        engine: str = "",
        sanitizer: str = "",
        architecture: str = "",
        root_cause_commit: str = "",
        root_cause_workspace: str = "",
        verify_top_3: bool = False

) -> dict:
    def normalize_report_text(text: str) -> str:
        return "\n".join(line.lstrip() for line in text.splitlines()).strip()

    def finalize_localization(status: str, detail_content: str) -> dict:
        artifact_path = "generated_prompt_file/commit_changed.txt"
        file_content = f"[STATUS]: {status}\n[TIMESTAMP]: {datetime.now().isoformat()}\n\n{detail_content}"
        # 强制更新工件文件
        create_or_update_file(artifact_path, file_content)
        # 如果是成功，返回结构化元数据；如果是失败，返回状态信息
        return {"status": status.lower(), "message": detail_content}

    if not ENABLE_HISTORY_ENHANCEMENT:
        detail_report = """[FAILURE_REGION]
(ECRCL ablated by configuration)

[ATTRIBUTION_TYPE]
UNKNOWN

[LOCALIZATION_CONFIDENCE]
LOW

[ROOT_CAUSE_COMMITS]
N/A

[CAUSAL_CHAIN]
Localization disabled by ablation switch.

[FINAL_ATTRIBUTION]
ECRCL disabled.
"""
        return finalize_localization("FAILED", detail_report)

    logger.info("=========================================================")
    logger.info(f"Starting Enhanced ECRCL Localization Engine for {project_name}")
    logger.info("=========================================================")

    print(f"[DEBUG] run_ecrcl_localization called with commit='{root_cause_commit}'")
    # 兼容环境变量获取底层编译依赖
    engine = engine or os.environ.get("ENGINE", "libfuzzer")
    sanitizer = sanitizer or os.environ.get("SANITIZER", "address")
    architecture = architecture or os.environ.get("ARCHITECTURE", "x86_64")

    if root_cause_commit and root_cause_workspace:
        logger.info(
            f"--- [ECRCL Fast-Path] Using pre-specified root cause for {project_name}: {root_cause_commit} in {root_cause_workspace} ---")
        is_downstream = (root_cause_workspace.upper() == "DOWNSTREAM")
        active_workspace = os.path.abspath(oss_fuzz_path) if is_downstream else os.path.abspath(project_source_path)

        log_artifact_path = os.path.abspath("fuzz_build_log_file/fuzz_build_log.txt")
        failure_region_text = "N/A"
        top_1_file = "N/A"
        line_number = "N/A"

        try:
            if os.path.exists(log_artifact_path):
                with open(log_artifact_path, 'r', encoding='utf-8', errors='ignore') as f:
                    log_lines = f.read().splitlines()
                tail_lines = log_lines[-40:] if log_lines else []
                if tail_lines:
                    failure_region_text = "\n".join(tail_lines)
                    matched_idx = -1
                    for i in range(len(tail_lines) - 1, -1, -1):
                        if any(kw in tail_lines[i].lower() for kw in ["error:", "cannot ", "fail", "undefined reference"]):
                            matched_idx = i
                            break
                    if matched_idx == -1:
                        for i in range(len(tail_lines) - 1, -1, -1):
                            if any(kw in tail_lines[i].lower() for kw in ["warning:", "exit status"]):
                                matched_idx = i
                                break
                    if matched_idx != -1:
                        start_idx = max(0, matched_idx - 10)
                        end_idx = min(len(tail_lines), matched_idx + 11)
                        failure_region_text = "\n".join(tail_lines[start_idx:end_idx])

                    path_pattern = r"([\w\-\./_]+\.(?:c|cpp|h|cc|cxx|rs|go|py|sh|java|swift|cmake|txt|yaml|json|PC|pc))"
                    raw_filepaths = re.findall(path_pattern, failure_region_text)
                    for f_cand in raw_filepaths:
                        if not any(sys_p in f_cand for sys_p in ["/usr/include/", "/.cargo/", "/.rustup/", "gcr.io/"]):
                            if f_cand.endswith(('.c', '.cpp', '.cc', '.h', '.go', '.rs', '.sh', 'Dockerfile', 'build.sh', 'PC', 'pc')):
                                top_1_file = f_cand
                                break
                    if top_1_file == "N/A" and raw_filepaths:
                        top_1_file = raw_filepaths[0]

                    line_match = re.search(rf"{re.escape(top_1_file)}:(\d+)", failure_region_text) if top_1_file != "N/A" else None
                    if line_match:
                        line_number = line_match.group(1)
        except Exception as e:
            logger.warning(f"Failed to recover failure evidence for fast-path localization: {e}")

        # 尝试提取元数据作为补全
        target_author, target_date, target_title = "N/A", "N/A", "N/A"
        diff_text = "Failed to extract commit diff context."
        before_line = "N/A"
        after_line = "N/A"

        try:
            show_meta = ["git", "-C", active_workspace, "show", "--pretty=format:%an|%ad|%s", "-s", root_cause_commit]
            meta_res = subprocess.run(show_meta, capture_output=True, text=True, check=True)
            target_author, target_date, target_title = meta_res.stdout.strip().split('|', 2)

            diff_res = subprocess.run(["git", "-C", active_workspace, "show", "-U3", root_cause_commit],
                                      capture_output=True, text=True, check=True)
            diff_text = clamp_diff_content(diff_res.stdout)
            removed_lines = [l[1:].strip() for l in diff_res.stdout.splitlines() if l.startswith('-') and not l.startswith('---')]
            added_lines = [l[1:].strip() for l in diff_res.stdout.splitlines() if l.startswith('+') and not l.startswith('+++')]
            if removed_lines:
                before_line = removed_lines[0]
            if added_lines:
                after_line = added_lines[0]

            if top_1_file == "N/A":
                for line in diff_res.stdout.splitlines():
                    diff_match = re.match(r"diff --git a/(.+?) b/", line)
                    if diff_match:
                        top_1_file = diff_match.group(1)
                        break
        except Exception as e:
            logger.warning(f"Failed to fetch pre-specified metadata: {e}")

            # 🔑 修正：账本回填移出 except 块，无条件执行，与元数据提取结果无关
        try:
            ledger = TraceLedgerManager.load_ledger()
            if ledger.get("nodes"):
                latest_node_id = ledger["nodes"][-1]["node_id"]
                TraceLedgerManager.update_node_fields(latest_node_id, {
                    "action_and_intent.root_cause_commit_sha": root_cause_commit
                })
                logger.info(
                    f"--- [Ledger] Pre-specified root cause SHA {root_cause_commit} written to Node {latest_node_id} ---")
        except Exception as e:
            logger.warning(f"Failed to auto-update pre-specified root_cause_commit_sha in ledger: {e}")

        detail_report = textwrap.dedent(f"""
        [FAILURE_REGION]
        {failure_region_text}

        [ATTRIBUTION_TYPE]
        {root_cause_workspace.upper()}

        [LOCALIZATION_CONFIDENCE]
        HIGH

        [ROOT_CAUSE_COMMITS]
        SHA: {root_cause_commit}
        Author: {target_author}
        Date: {target_date}
        Subject: {target_title}
        Reason: Pre-specified in metadata.

        [ROOT_CAUSE_LINES]
        File: {top_1_file}
        Line: {line_number}
        Before: {before_line}
        After: {after_line}

        [DIFF_CONTEXT]
        {diff_text}

        [CAUSAL_CHAIN]
        Pre-specified root cause bypassed candidate search and replay validation.
        Phase Four evidence was reconstructed from the current build log tail and commit diff.

        [FINAL_ATTRIBUTION]
        Root cause attribution completed from pre-specified commit metadata, log evidence, and diff context.
        """).strip()
        detail_report = normalize_report_text(detail_report)
        return finalize_localization("SUCCESS", detail_report)
    else:
        logger.info("--- [ECRCL] No pre-specified root cause. Starting Full Search ---")
        try:
            env_vars = dict(os.environ)

            # 0. 时间归一化
            t_error_epoch = timezone_normalize(error_date)
            t_error_utc = datetime.fromtimestamp(t_error_epoch, tz=timezone.utc)

            # 1. 故障日志提取 (Phase 0)
            if not os.path.exists(log_path):
                return finalize_localization("FAILED", f"Log file not found: {log_path}")

            with open(log_path, 'r', encoding='utf-8', errors='ignore') as f:
                log_raw = f.read()

            val_marker = "--- VALIDATION SUMMARY"
            raw_compile_zone = log_raw.split(val_marker)[0] if val_marker in log_raw else log_raw
            log_lines = raw_compile_zone.splitlines()

            matched_idx = -1
            for i in range(len(log_lines) - 1, -1, -1):
                if any(kw in log_lines[i].lower() for kw in ["error:", "cannot ", "fail", "undefined reference"]):
                    matched_idx = i
                    break

            if matched_idx == -1:
                for i in range(len(log_lines) - 1, -1, -1):
                    if any(kw in log_lines[i].lower() for kw in ["warning:", "exit status"]):
                        matched_idx = i
                        break

            if matched_idx == -1:
                return finalize_localization("FAILED", "No build failure features detected.")

            start_idx = max(0, matched_idx - 30)
            end_idx = min(len(log_lines), matched_idx + 31)
            failure_region_text = "\n".join(log_lines[start_idx:end_idx])

            # 提取区域内的相关代码或配置文件
            path_pattern = r"([\w\-\./_]+\.(?:c|cpp|h|cc|cxx|rs|go|py|sh|java|swift|cmake|txt|yaml|json|PC|pc))"
            raw_filepaths = re.findall(path_pattern, failure_region_text)
            line_number = "N/A"

            top_1_file = None
            for f_cand in raw_filepaths:
                if not any(sys_p in f_cand for sys_p in ["/usr/include/", "/.cargo/", "/.rustup/", "gcr.io/"]):
                    if f_cand.endswith(
                            ('.c', '.cpp', '.cc', '.h', '.go', '.rs', '.sh', 'Dockerfile', 'build.sh', 'PC', 'pc')):
                        top_1_file = f_cand
                        break
            if not top_1_file:
                top_1_file = raw_filepaths[0] if raw_filepaths else "build.sh"

            # 2. 确定初始搜索工作区 (Phase 1)
            is_downstream = any(
                cfg in top_1_file for cfg in ["Dockerfile", "build.sh", "project.yaml", "oss-fuzz", "projects/"])
            active_workspace = os.path.abspath(oss_fuzz_path) if is_downstream else os.path.abspath(project_source_path)

            logger.info(f"Targeting active workspace: {active_workspace} based on file: {top_1_file}")

            suspect_commits = []
            blamed_sha = None

            # Phase 1.1: 行级 Blame 追踪
            line_match = re.search(rf"{re.escape(top_1_file)}:(\d+)", failure_region_text)
            if line_match and os.path.exists(active_workspace):
                line_num = int(line_match.group(1))
                file_abs_path = os.path.abspath(os.path.join(active_workspace, top_1_file))
                if os.path.exists(file_abs_path) and os.path.isfile(file_abs_path):
                    try:
                        with open(file_abs_path, 'r', encoding='utf-8', errors='ignore') as f:
                            file_len = sum(1 for _ in f)
                        clamped_line = min(line_num, file_len) if file_len > 0 else 1
                        blame_cmd = ["git", "-C", active_workspace, "blame", "-L", f"{clamped_line},{clamped_line}",
                                     "--porcelain", file_abs_path]
                        res = subprocess.run(blame_cmd, capture_output=True, text=True, check=True)
                        blamed_sha = res.stdout.splitlines()[0].split(' ')[0]
                        logger.info(f"Precise Blame anchor matched: {blamed_sha}")
                    except Exception as e:
                        logger.warning(f"Git blame failed on precise line check: {e}")

            # Phase 1.3: 时域滑动窗口过滤 (T_error ± 24h)
            since_date = (t_error_utc - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
            until_date = (t_error_utc + timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')

            try:
                log_cmd = ["git", "-C", active_workspace, "log", f"--since={since_date}", f"--until={until_date}",
                           "--pretty=format:%H|%ct|%an|%cd|%s"]
                git_res = subprocess.run(log_cmd, capture_output=True, text=True, check=True)
                for line in git_res.stdout.splitlines():
                    if not line: continue
                    sha, epoch, author, date_str, msg = line.split('|', 4)
                    suspect_commits.append({
                        "sha": sha, "epoch": int(epoch), "author": author,
                        "date": date_str, "message": msg, "changed_files": []
                    })
            except Exception as e:
                logger.error(f"Failed to query git logs: {e}")

            # 兜底
            if not suspect_commits:
                try:
                    log_cmd = ["git", "-C", active_workspace, "log", "-n", "20", "--pretty=format:%H|%ct|%an|%cd|%s"]
                    git_res = subprocess.run(log_cmd, capture_output=True, text=True, check=True)
                    for line in git_res.stdout.splitlines():
                        if not line: continue
                        sha, epoch, author, date_str, msg = line.split('|', 4)
                        suspect_commits.append({
                            "sha": sha, "epoch": int(epoch), "author": author,
                            "date": date_str, "message": msg, "changed_files": []
                        })
                except Exception:
                    pass

            # 3. 约束收缩过滤器 (Phase 2)
            C1 = []
            for c in suspect_commits:
                try:
                    show_cmd = ["git", "-C", active_workspace, "show", "--name-only", "--format=", c["sha"]]
                    files_res = subprocess.run(show_cmd, capture_output=True, text=True, check=True)
                    c_files = [f.strip() for f in files_res.stdout.splitlines() if f.strip()]
                    c["changed_files"] = c_files

                    is_consistent = False
                    for f in c_files:
                        if os.path.basename(f) == os.path.basename(top_1_file):
                            is_consistent = True
                            break
                        if any(cfg in f for cfg in ["Dockerfile", "build.sh", "Makefile", "CMakeLists.txt"]):
                            is_consistent = True
                            break
                        if any(dep in f for dep in ["go.mod", "go.sum", "Cargo.toml", "package.json"]):
                            is_consistent = True
                            break
                    if is_consistent:
                        C1.append(c)
                except Exception:
                    pass

            if len(C1) >= 1:
                suspect_commits = C1

            # 4. 构建异构证据图谱并运行信念传播 (Phase 2.5)
            active_shas = [c["sha"] for c in suspect_commits]
            commit_messages_map = {c["sha"]: c["message"] for c in suspect_commits}

            graph = EvidenceGraph()
            graph.add_node("Nregion", "Failure Region")

            for c in suspect_commits:
                c_node = f"Ncommit_{c['sha']}"
                graph.add_node(c_node, "Commit")
                msg_node = f"Nmsg_{c['sha']}"
                graph.add_node(msg_node, "Commit Message")
                graph.add_edge(msg_node, c_node, 1.0)
                if any(kw in c["message"].lower() for kw in ["error", "fail", "conflict", "compile", "fix"]):
                    graph.add_edge("Nregion", msg_node, 1.1)
                for f in c["changed_files"]:
                    f_node = f"Nfile_{f}"
                    graph.add_node(f_node, "File")
                    dt = abs(c["epoch"] - t_error_epoch)
                    decay_weight = max(0.1, 1.5 * (0.5 ** (dt / 86400.0)))
                    graph.add_edge(f_node, c_node, decay_weight)
                    if os.path.basename(f) == os.path.basename(top_1_file):
                        graph.add_edge("Nregion", f_node, 1.3)
                if blamed_sha and c["sha"] == blamed_sha:
                    graph.add_edge("Nregion", c_node, 1.5)

            scores = graph.run_belief_propagation(active_shas, env_vars, commit_messages_map)
            sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)

            # =========================================================================
            # 🔑 新增合并核心逻辑： Phase 3 双轨反事实顺序因果校验状态机
            # =========================================================================
            if verify_top_3:
                suspect_pool = [score_info[0] for score_info in sorted_scores[:3]] if sorted_scores else []
                max_attempts = 3
            else:
                # 🔑 修正：当非 verify_top_3 时，直接通过下标解包首个元素，消除 score_info 未定义异常
                suspect_pool = [sorted_scores[0][0]] if sorted_scores else []
                max_attempts = 1

            final_suspect = "UNKNOWN"
            confidence = "LOW"
            validation_status = "FAIL"
            verification_passed = False

            for attempt_idx, suspect in enumerate(suspect_pool):
                logger.info(
                    f"--- [ECRCL Phase 3] Validation Attempt {attempt_idx + 1}/{max_attempts}: Testing {suspect} ---")
                try:
                    if is_downstream:
                        # [下游] 走安全回滚验证逻辑
                        revert_success = git_revert_counterfactual(
                            repo_path=active_workspace,
                            target_commit=suspect,
                            project_name=project_name,
                            engine=engine,
                            sanitizer=sanitizer,
                            architecture=architecture
                        )
                        parent_passed = revert_success
                        suspect_failed = True  # 原构建基线已知失败
                    else:
                        # [上游] 走双节点 checkout + compile 因果检验逻辑
                        # A. 物理暂存现场，避免切换分支遗失文件
                        subprocess.run(["git", "-C", active_workspace, "stash", "--include-untracked"],
                                       capture_output=True)
                        orig_head = subprocess.run(["git", "-C", active_workspace, "rev-parse", "HEAD"],
                                                   capture_output=True, text=True).stdout.strip()

                        # B. 检验父提交 suspect~1 (期望编译 SUCCESS)
                        checkout_parent = checkout_project_commit(active_workspace, f"{suspect}~1")
                        if checkout_parent["status"] == "success":
                            parent_passed = run_fuzz_build_and_validate(
                                project_name=project_name,
                                oss_fuzz_path=oss_fuzz_path,
                                sanitizer=sanitizer,
                                engine=engine,
                                architecture=architecture,
                                mount_path=project_source_path,
                                verbose_step6=False,
                                verbose_build=False
                            )["status"] == "success"
                        else:
                            parent_passed = False

                        # C. 检验当前提交 suspect (期望编译 FAIL)
                        checkout_suspect = checkout_project_commit(active_workspace, suspect)
                        if checkout_suspect["status"] == "success":
                            suspect_failed = run_fuzz_build_and_validate(
                                project_name=project_name,
                                oss_fuzz_path=oss_fuzz_path,
                                sanitizer=sanitizer,
                                engine=engine,
                                architecture=architecture,
                                mount_path=project_source_path,
                                verbose_step6=False,
                                verbose_build=False
                            )["status"] != "success"
                        else:
                            suspect_failed = False

                        # D. 强制复原上游仓库现场
                        checkout_project_commit(active_workspace, orig_head)
                        subprocess.run(["git", "-C", active_workspace, "clean", "-fdx"], capture_output=True)
                        subprocess.run(["git", "-C", active_workspace, "stash", "pop"], capture_output=True)

                    # 判定因果链：剔除代码能编过 且 保留代码编不过 => 判定该候选为因果根因
                    if parent_passed and suspect_failed:
                        validation_status = "PASS"
                        confidence = "HIGH"
                        final_suspect = suspect
                        verification_passed = True
                        logger.info(f"Causal Counterfactual validation PASSED on Attempt {attempt_idx + 1}!")

                        # 🔑 新增：当且仅当反事实检验判定成功时，物理触发账本归填
                        try:
                            ledger = TraceLedgerManager.load_ledger()
                            if ledger.get("nodes"):
                                latest_node_id = ledger["nodes"][-1]["node_id"]
                                TraceLedgerManager.update_node_fields(latest_node_id, {
                                    "action_and_intent.root_cause_commit_sha": final_suspect
                                })
                                logger.info(
                                    f"--- [Ledger] Discovered root cause SHA {final_suspect} written to Node {latest_node_id} ---")
                        except Exception as e:
                            logger.warning(f"Failed to auto-update discovered root_cause_commit_sha in ledger: {e}")
                        break
                    else:
                        logger.warning(f"Attempt {attempt_idx + 1} failed. Criteria not satisfied.")
                except Exception as val_err:
                    logger.error(f"Replay validation hit unexpected error on {suspect}: {val_err}")

            # =========================================================================
            # 🔑 新增合并核心逻辑： 物理元数据提取 (Metadata Extraction)
            # =========================================================================
            diff_text = ""
            target_author = "N/A"
            target_date = "N/A"
            target_title = "N/A"
            before_line = "N/A"
            after_line = "N/A"

            if final_suspect == "UNKNOWN":
                artifact_status = "FAILED"
                detail_report = textwrap.dedent(f"""
                [FAILURE_REGION]
                {failure_region_text}

                [ATTRIBUTION_TYPE]
                {'DOWNSTREAM' if is_downstream else 'UPSTREAM'}

                [LOCALIZATION_CONFIDENCE]
                LOW

                [ROOT_CAUSE_COMMITS]
                SHA: N/A
                Reason: Localization failed. System has automatically switched to Log-Based Diagnostic Mode.

                [ROOT_CAUSE_LINES]
                File: {top_1_file}
                Line: {line_number if 'line_number' in locals() else 'N/A'}
                Before: N/A
                After: N/A

                [DIFF_CONTEXT]
                N/A

                [CAUSAL_CHAIN]
                ECRCL localization failed, employing log-based diagnostic fallback.

                [FINAL_ATTRIBUTION]
                Root cause localization failed; proceed with log-based structural fix.
                """).strip()
            else:
                artifact_status = "SUCCESS"
                # 提取 Git 元数据逻辑 (保留您原有的 try-except)
                try:
                    show_meta = ["git", "-C", active_workspace, "show", "--pretty=format:%an|%ad|%s", "-s",
                                 final_suspect]
                    meta_res = subprocess.run(show_meta, capture_output=True, text=True, check=True)
                    target_author, target_date, target_title = meta_res.stdout.strip().split('|', 2)

                    diff_res = subprocess.run(["git", "-C", active_workspace, "show", "-U3", final_suspect],
                                              capture_output=True, text=True, check=True)
                    diff_text = clamp_diff_content(diff_res.stdout)
                    removed_lines = [l[1:].strip() for l in diff_res.stdout.splitlines() if
                                     l.startswith('-') and not l.startswith('---')]
                    added_lines = [l[1:].strip() for l in diff_res.stdout.splitlines() if
                                   l.startswith('+') and not l.startswith('+++')]
                    if removed_lines: before_line = removed_lines[0]
                    if added_lines: after_line = added_lines[0]
                except Exception:
                    diff_text = "Failed to extract commit diff context."

                detail_report = textwrap.dedent(f"""
                [FAILURE_REGION]
                {failure_region_text}

                [ATTRIBUTION_TYPE]
                {'DOWNSTREAM' if is_downstream else 'UPSTREAM'}

                [LOCALIZATION_CONFIDENCE]
                {confidence}

                [ROOT_CAUSE_COMMITS]
                SHA: {final_suspect}
                Author: {target_author}
                Date: {target_date}
                Subject: {target_title}

                [ROOT_CAUSE_LINES]
                File: {top_1_file}
                Line: {line_match.group(1) if line_match else 'N/A'}
                Before: {before_line}
                After: {after_line}

                [DIFF_CONTEXT]
                {diff_text}

                [CAUSAL_CHAIN]
                Replay verification status: {validation_status}.
                Belief propagation scores: {sorted_scores[:3]}.
                The root cause commit '{final_suspect}' introduces changes that broke compile check.

                [FINAL_ATTRIBUTION]
                Root cause localization successful.
                """).strip()

            detail_report = normalize_report_text(detail_report)

            return finalize_localization(artifact_status, detail_report)

        except Exception as e:

            logger.error(f"ECRCL localization critical error: {e}")
            return {"status": "failed", "message": f"Git localization error: {str(e)}"}
